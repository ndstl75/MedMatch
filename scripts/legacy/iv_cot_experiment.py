#!/usr/bin/env python3
"""
IV Chain-of-Thought (CoT) experiment for MedMatch.

Two-pass pipeline per medication order:
  Pass 1 – REASON: ask the LLM to think step-by-step about the order
           (identify drug, route, perform calculations, detect titratable).
  Pass 2 – EXTRACT: given the reasoning, extract into MedMatch JSON
           using structured JSON mode.

Covers all four IV sub-categories:
  - IV intermittent  (n=16)
  - IV push          (n=17)
  - IV continuous titratable     (n=11)
  - IV continuous non-titratable (n=6)

Scoring reports titratable and non-titratable separately.
"""

import csv
import json
import os
import re
import sys
import time
from datetime import datetime

import openpyxl
from google import genai
from google.genai import errors, types


ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, "src")
if SRC not in sys.path:
    sys.path.insert(0, SRC)

# ── Config ────────────────────────────────────────────────────────────────────
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
if not GOOGLE_API_KEY and os.path.exists(os.path.join(ROOT, ".env")):
    with open(os.path.join(ROOT, ".env"), "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("GOOGLE_API_KEY="):
                GOOGLE_API_KEY = line.strip().split("=", 1)[1]
                break

if not GOOGLE_API_KEY:
    raise RuntimeError("GOOGLE_API_KEY is not set. Add it to your environment or .env.")

CLIENT = genai.Client(api_key=GOOGLE_API_KEY)
MODEL_NAME = os.environ.get("GOOGLE_MODEL_NAME", "gemma-3-27b-it")
TEMPERATURE = float(os.environ.get("MEDMATCH_TEMPERATURE", "0.1"))
NUM_RUNS = int(os.environ.get("MEDMATCH_NUM_RUNS", "3"))  # triplicate per paper
MAX_RETRIES = 3
RETRY_DELAY = float(os.environ.get("MEDMATCH_RETRY_DELAY", "5"))
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "1"))
SELECTED_SHEETS = [
    s.strip() for s in os.environ.get("MEDMATCH_SHEETS", "").split(",") if s.strip()
]
MAX_ENTRIES_PER_SHEET = int(os.environ.get("MEDMATCH_MAX_ENTRIES", "0"))

SYSTEM_PROMPT = (
    "You are a clinical pharmacist who formats medication orders. "
    "Only output the MedMatch JSON format."
)

KEY_ALIASES = {
    "rate": "starting rate",
    "volume amount": "numerical volume",
    "amount of diluent": "amount of diluent volume",
    "diluent amount": "amount of diluent volume",
    "volume unit": "volume unit of measure",
    "dose unit": "abbreviated unit strength of dose",
}


# ── Prompts ───────────────────────────────────────────────────────────────────

# Import base instructions from the shared package layer.
from medmatch.core.schema import (
    IV_INTERMITTENT_INSTRUCTION,
    IV_PUSH_INSTRUCTION,
    IV_CONTINUOUS_INSTRUCTION,
)


# Pass 1: Reasoning prompt — ask the model to think step-by-step BEFORE
# extracting.  The reasoning is free-form text (no JSON mode).
COT_REASON_PROMPT = {
    "IV intermittent (16)": """\
You are a clinical pharmacist. Analyze the following IV intermittent medication order step by step.

For each step, state what you find:
1. Drug name: identify the generic or brand name (lowercase).
2. Dose: the numeric drug dose and its unit (e.g., mg, g, units).
3. Diluent: the volume of diluent (numeric, in mL) and the compatible diluent type (e.g., 0.9% sodium chloride, D5W).
4. Infusion time: how long the infusion runs (e.g., 30 minutes, 1 hour).
5. Frequency: how often the dose is given (e.g., every 8 hours, once daily). If the order says just "daily", interpret as "once daily".

Medication order:
{prompt}

Think through each step, then summarize your findings clearly.""",

    "IV push (17)": """\
You are a clinical pharmacist. Analyze the following IV push medication order step by step.

For each step, state what you find:
1. Drug name: identify the generic or brand name (lowercase).
2. Dose: the numeric drug dose and its unit (e.g., mg, mcg, units).
3. Volume: the numeric volume administered (in mL).
4. Concentration: calculate the concentration per 1 mL. If the order says "20 mg/2 mL", the concentration is 10 mg/mL. If "50 mcg/mL", it is 50 mcg/mL. Always express as [number] [unit]/mL.
5. Formulation: the dosage form — always "vial solution" if the order mentions a vial.
6. Frequency: how often the push is given. If the order says just "daily", interpret as "once daily".

Medication order:
{prompt}

Think through each step, then summarize your findings clearly.""",

    "IV continuous (16)": """\
You are a clinical pharmacist. Analyze the following IV continuous infusion medication order step by step.

For each step, state what you find:
1. Drug name: identify the generic or brand name (lowercase).
2. Total dose in the bag: This is the TOTAL amount of drug in the prepared bag, NOT the per-mL concentration.
   - If the order says "4 mg/250 mL", the total dose is 4 mg.
   - If the order says "1 mg/mL in 100 mL", the total dose is 1 × 100 = 100 mg.
   - If the order says "10 mg/mL concentration" with no volume, try to infer the total from context.
   State the numeric dose and its unit (mg, mcg, units, etc.).
3. Diluent volume: the total volume of the bag in mL.
4. Compatible diluent type: the IV fluid (e.g., 0.9% sodium chloride, D5W).
5. Starting rate: the initial infusion rate and its unit (e.g., 0.05 mcg/kg/min, 5 mg/hr).
   - Normalize "hour" to "hr" in rate units (e.g., mg/hr not mg/hour).
6. Titratable or non-titratable?
   - If the order mentions titration, titrate, adjust, increase/decrease by, or a clinical goal → TITRATABLE.
   - Otherwise → NON-TITRATABLE. Leave titration fields empty.
7. If titratable:
   a. Titration dose: the amount the rate changes per step.
   b. Titration unit: the unit of the titration increment.
   c. Titration frequency: how often titration can occur (e.g., every 5 minutes).
   d. Titration goal: copy the clinical target from the order as closely as possible. Do NOT paraphrase.

Medication order:
{prompt}

Think through each step, then summarize your findings clearly.""",
}


# Pass 2: Extraction prompt — given the reasoning, produce MedMatch JSON.
COT_EXTRACT_PREFIX = """\
Based on the clinical analysis below, extract the medication information into the MedMatch JSON format.

ANALYSIS:
{reasoning}

ORIGINAL ORDER:
{prompt}

INSTRUCTIONS:
{base_instruction}

Return one JSON object only.
Do not wrap the JSON in markdown.
Use exactly these keys in this order: {keys}.
{extra_guidance}
Now produce the MedMatch JSON:"""


# Extra guidance per sheet to handle common edge cases
EXTRACT_GUIDANCE = {
    "IV intermittent (16)": (
        'If frequency is just "daily", write "once daily".'
    ),
    "IV push (17)": (
        "For concentration: always convert to per-1-mL basis "
        "(e.g., 20 mg/2 mL → concentration of solution = 10, "
        "concentration unit of measure = mg/mL). "
        'Formulation should be "vial solution" when a vial is mentioned. '
        'If frequency is just "daily", write "once daily".'
    ),
    "IV continuous (16)": (
        "Use the TOTAL drug amount in the bag for numerical dose, "
        "not the per-mL concentration. "
        "Normalize hour to hr in rate units (mg/hr, mcg/hr, units/hr). "
        "For non-titratable infusions, leave titration fields as empty strings. "
        "Copy the titration goal text closely from the original order."
    ),
}


# ── Sheet config ──────────────────────────────────────────────────────────────

SHEET_CONFIG = {
    "IV intermittent (16)": {
        "instruction": IV_INTERMITTENT_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "infusion time": 10,
            "frequency": 11,
        },
    },
    "IV push (17)": {
        "instruction": IV_PUSH_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of volume": 7,
            "volume unit of measure": 8,
            "concentration of solution": 9,
            "concentration unit of measure": 10,
            "formulation": 11,
            "frequency": 12,
        },
    },
    "IV continuous (16)": {
        "instruction": IV_CONTINUOUS_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "starting rate": 10,
            "unit of measure": 11,
            "titration dose": 12,
            "titration unit of measure": 13,
            "titration frequency": 14,
            "titration goal based on physiologic response, laboratory result, or assessment score": 15,
        },
    },
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_key(key):
    s = str(key).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return KEY_ALIASES.get(s, s)


def normalize(value):
    """Normalize a value for comparison."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = s.replace("–", "-").replace("—", "-").replace("≥", ">=")
    s = re.sub(r"\bmg/hour\b", "mg/hr", s)
    s = re.sub(r"\bmcg/hour\b", "mcg/hr", s)
    s = re.sub(r"\bunits/hour\b", "units/hr", s)
    s = re.sub(r"\bunit/hour\b", "units/hr", s)
    s = re.sub(r"\s*mmhg\b", " mmhg", s)
    s = re.sub(r"\s+", " ", s)
    try:
        f = float(s)
        s = str(int(f)) if f == int(f) else str(f)
    except (ValueError, OverflowError):
        pass
    return s


def compare_results(llm_output, ground_truth):
    results = {}
    for key, expected in ground_truth.items():
        e = normalize(expected)
        a = normalize(llm_output.get(key, "") if llm_output else "")
        results[key] = {"expected": e, "actual": a, "match": e == a}
    return results


def all_fields_match(comp):
    return all(v["match"] for v in comp.values())


def is_titratable(ground_truth):
    """Determine if a continuous entry is titratable from its ground truth."""
    tit_fields = [
        "titration dose",
        "titration unit of measure",
        "titration frequency",
        "titration goal based on physiologic response, laboratory result, or assessment score",
    ]
    return any(
        ground_truth.get(f) not in (None, "", 0)
        for f in tit_fields
    )


def build_response_schema(expected_keys):
    return {
        "type": "object",
        "properties": {
            key: {"type": ["string", "number", "integer", "null"]}
            for key in expected_keys
        },
        "required": expected_keys,
        "additionalProperties": False,
    }


def parse_json_response(text):
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3].strip()
        elif "```" in text:
            text = text[: text.rfind("```")].strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            start = text.index("{")
            end = text.rindex("}") + 1
            return json.loads(text[start:end])
        except (ValueError, json.JSONDecodeError):
            return None


def coerce_output_object(parsed, expected_keys):
    if parsed is None:
        return {key: "" for key in expected_keys}
    if isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        return {key: "" for key in expected_keys}
    normalized = {}
    for key, value in parsed.items():
        normalized[normalize_key(key)] = value
    return {key: normalized.get(key, "") for key in expected_keys}


# ── API calls ─────────────────────────────────────────────────────────────────

def call_reason(prompt_text, retries=MAX_RETRIES):
    """Pass 1: free-form reasoning (no JSON mode)."""
    for attempt in range(retries):
        try:
            try:
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=prompt_text,
                    config=types.GenerateContentConfig(
                        system_instruction=(
                            "You are a clinical pharmacist. "
                            "Think step by step and be precise with numbers and units."
                        ),
                        temperature=TEMPERATURE,
                    ),
                )
            except errors.ClientError as e:
                if "Developer instruction is not enabled" not in str(e):
                    raise
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=prompt_text,
                    config=types.GenerateContentConfig(temperature=TEMPERATURE),
                )
            return response.text.strip()
        except Exception as e:
            if attempt < retries - 1:
                print(f" [reason retry {attempt+1}: {e}]", end="", flush=True)
                time.sleep(RETRY_DELAY)
            else:
                return ""


def call_extract(prompt_text, expected_keys, retries=MAX_RETRIES):
    """Pass 2: structured JSON extraction."""
    schema = build_response_schema(expected_keys)
    for attempt in range(retries):
        text = ""
        try:
            try:
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=prompt_text,
                    config=types.GenerateContentConfig(
                        system_instruction=SYSTEM_PROMPT,
                        temperature=TEMPERATURE,
                        response_mime_type="application/json",
                        response_json_schema=schema,
                    ),
                )
            except errors.ClientError as e:
                err = str(e)
                if "JSON mode is not enabled" not in err and "Developer instruction is not enabled" not in err:
                    raise
                fallback = f"{SYSTEM_PROMPT}\n\n{prompt_text}"
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=fallback,
                    config=types.GenerateContentConfig(temperature=TEMPERATURE),
                )
            text = response.text.strip()
            parsed = parse_json_response(text)
            return coerce_output_object(parsed, expected_keys), text
        except Exception as e:
            if attempt < retries - 1:
                print(f" [extract retry {attempt+1}: {e}]", end="", flush=True)
                time.sleep(RETRY_DELAY)
            else:
                return {k: "" for k in expected_keys}, text or str(e)


# ── Dataset loading ───────────────────────────────────────────────────────────

def load_dataset(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    for sheet_name, config in SHEET_CONFIG.items():
        if SELECTED_SHEETS and sheet_name not in SELECTED_SHEETS:
            continue
        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[config["prompt_col"] - 1].value
            if not prompt:
                break
            ground_truth = {}
            for key, col in config["ground_truth_cols"].items():
                val = row[col - 1].value
                ground_truth[key] = val if val is not None else ""
            entries.append({
                "medication": row[0].value,
                "prompt": str(prompt).strip(),
                "ground_truth": ground_truth,
            })
            if MAX_ENTRIES_PER_SHEET and len(entries) >= MAX_ENTRIES_PER_SHEET:
                break
        dataset[sheet_name] = entries
        print(f"  Loaded {len(entries)} entries from '{sheet_name}'")
    return dataset


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    xlsx_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "MedMatch Dataset for Experiment_ Final.xlsx",
    )
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
    os.makedirs(results_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"Model: {MODEL_NAME} | Temp: {TEMPERATURE} | Runs: {NUM_RUNS}")
    print(f"Pipeline: reason (CoT) → extract (JSON mode) → score")
    print(f"Loading dataset...")
    dataset = load_dataset(xlsx_path)

    grand_total = {"entries": 0, "correct": 0, "fields": 0, "fields_ok": 0}
    # Separate tracking for titratable vs non-titratable
    tit_stats = {"entries": 0, "correct": 0, "fields": 0, "fields_ok": 0}
    nontit_stats = {"entries": 0, "correct": 0, "fields": 0, "fields_ok": 0}

    for sheet_name, entries in dataset.items():
        config = SHEET_CONFIG[sheet_name]
        expected_keys = list(config["ground_truth_cols"].keys())
        reason_template = COT_REASON_PROMPT.get(sheet_name, "")
        guidance = EXTRACT_GUIDANCE.get(sheet_name, "")

        print(f"\n{'='*72}")
        print(f"{sheet_name} ({len(entries)} entries x {NUM_RUNS} runs)")
        print(f"{'='*72}")

        all_rows = []

        for run_idx in range(NUM_RUNS):
            run_ok = 0
            print(f"\n  --- Run {run_idx + 1}/{NUM_RUNS} ---")

            for idx, entry in enumerate(entries, 1):
                print(f"  [{idx}/{len(entries)}] {entry['medication']}...", end=" ", flush=True)

                # ── Pass 1: Reason ──
                reason_prompt = reason_template.format(prompt=entry["prompt"])
                reasoning = call_reason(reason_prompt)
                time.sleep(SLEEP_BETWEEN_CALLS)

                # ── Pass 2: Extract ──
                extract_prompt = COT_EXTRACT_PREFIX.format(
                    reasoning=reasoning,
                    prompt=entry["prompt"],
                    base_instruction=config["instruction"],
                    keys=", ".join(expected_keys),
                    extra_guidance=guidance,
                )
                llm_output, raw_extract = call_extract(extract_prompt, expected_keys)
                time.sleep(SLEEP_BETWEEN_CALLS)

                # ── Score ──
                comparison = compare_results(llm_output, entry["ground_truth"])
                n_match = sum(1 for v in comparison.values() if v["match"])
                n_total = len(comparison)
                entry_correct = all_fields_match(comparison)

                if entry_correct:
                    run_ok += 1

                # Track stats
                grand_total["entries"] += 1
                grand_total["correct"] += int(entry_correct)
                grand_total["fields"] += n_total
                grand_total["fields_ok"] += n_match

                # Split titratable vs non-titratable for IV continuous
                entry_type = None
                if sheet_name == "IV continuous (16)":
                    if is_titratable(entry["ground_truth"]):
                        entry_type = "titratable"
                        tit_stats["entries"] += 1
                        tit_stats["correct"] += int(entry_correct)
                        tit_stats["fields"] += n_total
                        tit_stats["fields_ok"] += n_match
                    else:
                        entry_type = "non-titratable"
                        nontit_stats["entries"] += 1
                        nontit_stats["correct"] += int(entry_correct)
                        nontit_stats["fields"] += n_total
                        nontit_stats["fields_ok"] += n_match

                status = f"{n_match}/{n_total}"
                if entry_correct:
                    status += " [ALL]"
                if entry_type:
                    status += f" ({entry_type})"
                print(status)

                all_rows.append({
                    "run": run_idx + 1,
                    "medication": entry["medication"],
                    "prompt": entry["prompt"],
                    "ground_truth": entry["ground_truth"],
                    "reasoning": reasoning,
                    "llm_output": llm_output,
                    "raw_extract_response": raw_extract,
                    "comparison": comparison,
                    "fields_correct": n_match,
                    "fields_total": n_total,
                    "all_fields_correct": entry_correct,
                    "entry_type": entry_type,
                })

            pct = run_ok / len(entries) * 100 if entries else 0
            print(f"\n  Run {run_idx + 1}: {run_ok}/{len(entries)} = {pct:.1f}%")

        # ── Save results for this sheet ──
        safe = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        json_path = os.path.join(results_dir, f"{safe}_cot_{timestamp}.json")
        csv_path = os.path.join(results_dir, f"{safe}_cot_{timestamp}.csv")

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, indent=2, default=str)

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header = ["Run", "Medication", "Type"] + expected_keys + [
                "Fields Correct", "All Correct"
            ]
            writer.writerow(header)
            for r in all_rows:
                row = [r["run"], r["medication"], r.get("entry_type", "")]
                for key in expected_keys:
                    c = r["comparison"].get(key, {})
                    if c.get("match"):
                        row.append("MATCH")
                    else:
                        row.append(
                            f"MISS (exp: {c.get('expected','')}, got: {c.get('actual','')})"
                        )
                row.append(f"{r['fields_correct']}/{r['fields_total']}")
                row.append("YES" if r["all_fields_correct"] else "NO")
                writer.writerow(row)

        print(f"\n  Saved: {json_path}")
        print(f"  Saved: {csv_path}")

        # Per-sheet summary
        sheet_rows = all_rows
        sheet_total = len(sheet_rows)
        sheet_ok = sum(1 for r in sheet_rows if r["all_fields_correct"])
        sheet_fc = sum(r["fields_correct"] for r in sheet_rows)
        sheet_ft = sum(r["fields_total"] for r in sheet_rows)
        print(f"  Sheet overall: {sheet_ok}/{sheet_total} = {sheet_ok/sheet_total*100:.1f}%")
        print(f"  Sheet field:   {sheet_fc}/{sheet_ft} = {sheet_fc/sheet_ft*100:.1f}%")

        # Split summary for IV continuous
        if sheet_name == "IV continuous (16)":
            tit_rows = [r for r in sheet_rows if r.get("entry_type") == "titratable"]
            nontit_rows = [r for r in sheet_rows if r.get("entry_type") == "non-titratable"]
            if tit_rows:
                tok = sum(1 for r in tit_rows if r["all_fields_correct"])
                print(f"  Titratable:     {tok}/{len(tit_rows)} = {tok/len(tit_rows)*100:.1f}%")
            if nontit_rows:
                ntok = sum(1 for r in nontit_rows if r["all_fields_correct"])
                print(f"  Non-titratable: {ntok}/{len(nontit_rows)} = {ntok/len(nontit_rows)*100:.1f}%")

    # ── Grand summary ──
    g = grand_total
    print(f"\n{'='*72}")
    print(f"GRAND SUMMARY (CoT pipeline, {MODEL_NAME}, {NUM_RUNS} runs)")
    print(f"  Overall accuracy: {g['correct']}/{g['entries']} = "
          f"{g['correct']/g['entries']*100:.1f}%")
    print(f"  Field accuracy:   {g['fields_ok']}/{g['fields']} = "
          f"{g['fields_ok']/g['fields']*100:.1f}%")
    if tit_stats["entries"]:
        t = tit_stats
        print(f"  IV CI titratable:     {t['correct']}/{t['entries']} = "
              f"{t['correct']/t['entries']*100:.1f}%")
    if nontit_stats["entries"]:
        n = nontit_stats
        print(f"  IV CI non-titratable: {n['correct']}/{n['entries']} = "
              f"{n['correct']/n['entries']*100:.1f}%")
    print(f"{'='*72}")


if __name__ == "__main__":
    main()
