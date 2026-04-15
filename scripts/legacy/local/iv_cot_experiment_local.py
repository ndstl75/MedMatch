#!/usr/bin/env python3
"""
IV Chain-of-Thought (CoT) experiment for MedMatch — LOCAL Ollama version.

Two-pass pipeline per medication order:
  Pass 1 – REASON: ask the LLM to think step-by-step about the order
           (identify drug, dose, volume, rate, titratable flag, etc.).
  Pass 2 – EXTRACT: given the reasoning, extract into MedMatch JSON
           using JSON format constraint.

Appendix-aligned: reason prompts name the category (IV intermittent / IV push
/ IV continuous) and walk through the appendix's declared fields, but do NOT
embed worked examples or normalization rules. Baseline appendix instructions
(IV_*_INSTRUCTION) supply format; the reason pass adds structured thinking only.

Covers all four IV sub-categories:
  - IV intermittent  (n=16)
  - IV push          (n=17)
  - IV continuous titratable     (n=11)
  - IV continuous non-titratable (n=6)

Scoring reports titratable and non-titratable separately.
"""

import csv
import json
import os
import re
import time
from datetime import datetime

from local_llm import chat_completion, resolve_project_file, OLLAMA_MODEL
from medmatch_test_local import (
    IV_CONTINUOUS_INSTRUCTION,
    IV_INTERMITTENT_INSTRUCTION,
    IV_PUSH_INSTRUCTION,
)


# ── Config ────────────────────────────────────────────────────────────────────
MODEL_NAME = os.environ.get("OLLAMA_MODEL", OLLAMA_MODEL)
TEMPERATURE = float(os.environ.get("MEDMATCH_TEMPERATURE", "0.1"))
NUM_RUNS = int(os.environ.get("MEDMATCH_NUM_RUNS", "3"))  # triplicate per paper
MAX_RETRIES = int(os.environ.get("MEDMATCH_MAX_RETRIES", "3"))
RETRY_DELAY = float(os.environ.get("MEDMATCH_RETRY_DELAY", "5"))
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "0.5"))
SELECTED_SHEETS = [
    s.strip() for s in os.environ.get("MEDMATCH_SHEETS", "").split(",") if s.strip()
]
MAX_ENTRIES_PER_SHEET = int(os.environ.get("MEDMATCH_MAX_ENTRIES", "0"))

SYSTEM_PROMPT = (
    "You are a clinical pharmacist who formats medication orders. "
    "Only output the MedMatch JSON format."
)

# ── Reasoning prompts (Pass 1) ───────────────────────────────────────────────

COT_REASON_PROMPT = {
    "IV intermittent (16)": """\
You are a clinical pharmacist. Analyze the following IV intermittent medication order step by step.

For each step, state what you find:
1. Drug name.
2. Dose (numeric value and unit).
3. Diluent volume and compatible diluent type.
4. Infusion time.
5. Frequency.

Medication order:
{prompt}

Think through each step, then summarize your findings clearly.""",

    "IV push (17)": """\
You are a clinical pharmacist. Analyze the following IV push medication order step by step.

For each step, state what you find:
1. Drug name.
2. Dose (numeric value and unit).
3. Volume administered.
4. Concentration of solution.
5. Formulation in canonical MedMatch wording.
6. Frequency in canonical MedMatch wording.

Medication order:
{prompt}

Think through each step, then summarize your findings clearly.""",

    "IV continuous (16)": """\
You are a clinical pharmacist. Analyze the following IV continuous infusion medication order step by step.

For each step, state what you find:
1. Drug name.
2. Drug dose in the prepared infusion (numeric value and unit).
3. Diluent volume.
4. Compatible diluent type.
5. Starting rate (value and unit).
6. Is this order titratable or non-titratable, based on the order text?
7. If titratable: titration dose, titration unit, titration frequency, and titration goal.

Medication order:
{prompt}

Think through each step, then summarize your findings clearly.""",
}


# ── Extraction prompt template (Pass 2) ──────────────────────────────────────

COT_EXTRACT_PREFIX = """\
Based on the clinical analysis below, extract the medication information into the MedMatch JSON format.

ANALYSIS:
{reasoning}

ORIGINAL ORDER:
{prompt}

INSTRUCTIONS:
{base_instruction}

Return one JSON object only.
Do not wrap the JSON in markdown.
Use exactly these keys in this order: {keys}.
Now produce the MedMatch JSON:"""


COT_EXTRACT_CANONICALIZATION = {
    "IV push (17)": """\
Canonicalization for IV push:
- If the source says vial or vial solution, output formulation as "vial solution".
- Normalize frequency wording to MedMatch canonical form:
  - daily -> once daily
  - BID -> twice daily""",
}


# ── Sheet config ──────────────────────────────────────────────────────────────

SHEET_CONFIG = {
    "IV intermittent (16)": {
        "instruction": IV_INTERMITTENT_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "infusion time": 10,
            "frequency": 11,
        },
    },
    "IV push (17)": {
        "instruction": IV_PUSH_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of volume": 7,
            "volume unit of measure": 8,
            "concentration of solution": 9,
            "concentration unit of measure": 10,
            "formulation": 11,
            "frequency": 12,
        },
    },
    "IV continuous (16)": {
        "instruction": IV_CONTINUOUS_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "starting rate": 10,
            "unit of measure": 11,
            "titration dose": 12,
            "titration unit of measure": 13,
            "titration frequency": 14,
            "titration goal based on physiologic response, laboratory result, or assessment score": 15,
        },
    },
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_key(key):
    """Strict: lowercase + whitespace-collapse only. No key aliases."""
    s = str(key).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize(value):
    """Strict: lowercase + whitespace-collapse only. No unit rewrites, no
    unicode substitution, no numeric coercion. Local-side normalization
    beyond case/whitespace would relax the string-exact match standard."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def compare_results(llm_output, ground_truth):
    results = {}
    for key, expected in ground_truth.items():
        e = normalize(expected)
        a = normalize(llm_output.get(key, "") if llm_output else "")
        results[key] = {"expected": e, "actual": a, "match": e == a}
    return results


def all_fields_match(comp):
    return all(v["match"] for v in comp.values())


def is_titratable(ground_truth):
    tit_fields = [
        "titration dose", "titration unit of measure", "titration frequency",
        "titration goal based on physiologic response, laboratory result, or assessment score",
    ]
    return any(ground_truth.get(f) not in (None, "", 0) for f in tit_fields)


def parse_json_response(text):
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3].strip()
        elif "```" in text:
            text = text[: text.rfind("```")].strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            start = text.index("{")
            end = text.rindex("}") + 1
            return json.loads(text[start:end])
        except (ValueError, json.JSONDecodeError):
            return None


def coerce_output_object(parsed, expected_keys):
    if parsed is None:
        return {key: "" for key in expected_keys}
    if isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        return {key: "" for key in expected_keys}
    normalized = {}
    for key, value in parsed.items():
        normalized[normalize_key(key)] = value
    return {key: normalized.get(key, "") for key in expected_keys}


# ── API calls (local Ollama) ─────────────────────────────────────────────────

def call_reason(prompt_text, retries=MAX_RETRIES):
    """Pass 1: free-form reasoning."""
    system = (
        "You are a clinical pharmacist. "
        "Think step by step and be precise with numbers and units."
    )
    for attempt in range(retries):
        try:
            return chat_completion(system, prompt_text, temperature=TEMPERATURE)
        except Exception as e:
            if attempt < retries - 1:
                print(f" [reason retry {attempt+1}: {e}]", end="", flush=True)
                time.sleep(RETRY_DELAY)
            else:
                return ""


def call_extract(prompt_text, expected_keys, retries=MAX_RETRIES):
    """Pass 2: JSON extraction."""
    for attempt in range(retries):
        text = ""
        try:
            text = chat_completion(SYSTEM_PROMPT, prompt_text, temperature=TEMPERATURE)
            parsed = parse_json_response(text)
            return coerce_output_object(parsed, expected_keys), text
        except Exception as e:
            if attempt < retries - 1:
                print(f" [extract retry {attempt+1}: {e}]", end="", flush=True)
                time.sleep(RETRY_DELAY)
            else:
                return {k: "" for k in expected_keys}, text or str(e)


# ── Dataset loading ───────────────────────────────────────────────────────────

def load_dataset(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    for sheet_name, config in SHEET_CONFIG.items():
        if SELECTED_SHEETS and sheet_name not in SELECTED_SHEETS:
            continue
        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[config["prompt_col"] - 1].value
            if not prompt:
                break
            ground_truth = {}
            for key, col in config["ground_truth_cols"].items():
                val = row[col - 1].value
                ground_truth[key] = val if val is not None else ""
            entries.append({
                "medication": row[0].value,
                "prompt": str(prompt).strip(),
                "ground_truth": ground_truth,
            })
            if MAX_ENTRIES_PER_SHEET and len(entries) >= MAX_ENTRIES_PER_SHEET:
                break
        dataset[sheet_name] = entries
        print(f"  Loaded {len(entries)} entries from '{sheet_name}'")
    return dataset


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    xlsx_path = resolve_project_file("MedMatch Dataset for Experiment_ Final.xlsx")
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
    os.makedirs(results_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"Model: {MODEL_NAME} (local Ollama) | Temp: {TEMPERATURE} | Runs: {NUM_RUNS}")
    print(f"Pipeline: reason (CoT) → extract (JSON) → score")
    print(f"Loading dataset...")
    dataset = load_dataset(xlsx_path)

    grand_total = {"entries": 0, "correct": 0, "fields": 0, "fields_ok": 0}
    tit_stats = {"entries": 0, "correct": 0, "fields": 0, "fields_ok": 0}
    nontit_stats = {"entries": 0, "correct": 0, "fields": 0, "fields_ok": 0}

    for sheet_name, entries in dataset.items():
        config = SHEET_CONFIG[sheet_name]
        expected_keys = list(config["ground_truth_cols"].keys())
        reason_template = COT_REASON_PROMPT.get(sheet_name, "")

        print(f"\n{'='*72}")
        print(f"{sheet_name} ({len(entries)} entries x {NUM_RUNS} runs)")
        print(f"{'='*72}")

        all_rows = []

        for run_idx in range(NUM_RUNS):
            run_ok = 0
            print(f"\n  --- Run {run_idx + 1}/{NUM_RUNS} ---")

            for idx, entry in enumerate(entries, 1):
                print(f"  [{idx}/{len(entries)}] {entry['medication']}...", end=" ", flush=True)

                # ── Pass 1: Reason ──
                reason_prompt = reason_template.format(prompt=entry["prompt"])
                reasoning = call_reason(reason_prompt)
                time.sleep(SLEEP_BETWEEN_CALLS)

                # ── Pass 2: Extract ──
                extract_prompt = COT_EXTRACT_PREFIX.format(
                    reasoning=reasoning,
                    prompt=entry["prompt"],
                    base_instruction="\n\n".join(
                        part
                        for part in (
                            config["instruction"],
                            COT_EXTRACT_CANONICALIZATION.get(sheet_name, ""),
                        )
                        if part
                    ),
                    keys=", ".join(expected_keys),
                )
                llm_output, raw_extract = call_extract(extract_prompt, expected_keys)
                time.sleep(SLEEP_BETWEEN_CALLS)

                # ── Score ──
                comparison = compare_results(llm_output, entry["ground_truth"])
                n_match = sum(1 for v in comparison.values() if v["match"])
                n_total = len(comparison)
                entry_correct = all_fields_match(comparison)

                if entry_correct:
                    run_ok += 1

                grand_total["entries"] += 1
                grand_total["correct"] += int(entry_correct)
                grand_total["fields"] += n_total
                grand_total["fields_ok"] += n_match

                entry_type = None
                if sheet_name == "IV continuous (16)":
                    if is_titratable(entry["ground_truth"]):
                        entry_type = "titratable"
                        tit_stats["entries"] += 1
                        tit_stats["correct"] += int(entry_correct)
                        tit_stats["fields"] += n_total
                        tit_stats["fields_ok"] += n_match
                    else:
                        entry_type = "non-titratable"
                        nontit_stats["entries"] += 1
                        nontit_stats["correct"] += int(entry_correct)
                        nontit_stats["fields"] += n_total
                        nontit_stats["fields_ok"] += n_match

                status = f"{n_match}/{n_total}"
                if entry_correct:
                    status += " [ALL]"
                if entry_type:
                    status += f" ({entry_type})"
                print(status)

                all_rows.append({
                    "run": run_idx + 1,
                    "medication": entry["medication"],
                    "prompt": entry["prompt"],
                    "ground_truth": entry["ground_truth"],
                    "reasoning": reasoning,
                    "llm_output": llm_output,
                    "raw_extract_response": raw_extract,
                    "comparison": comparison,
                    "fields_correct": n_match,
                    "fields_total": n_total,
                    "all_fields_correct": entry_correct,
                    "entry_type": entry_type,
                })

            pct = run_ok / len(entries) * 100 if entries else 0
            print(f"\n  Run {run_idx + 1}: {run_ok}/{len(entries)} = {pct:.1f}%")

        # ── Save results ──
        safe = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        json_path = os.path.join(results_dir, f"{safe}_cot_local_{timestamp}.json")
        csv_path = os.path.join(results_dir, f"{safe}_cot_local_{timestamp}.csv")

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, indent=2, default=str)

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header = ["Run", "Medication", "Type"] + expected_keys + [
                "Fields Correct", "All Correct"
            ]
            writer.writerow(header)
            for r in all_rows:
                row = [r["run"], r["medication"], r.get("entry_type", "")]
                for key in expected_keys:
                    c = r["comparison"].get(key, {})
                    if c.get("match"):
                        row.append("MATCH")
                    else:
                        row.append(
                            f"MISS (exp: {c.get('expected','')}, got: {c.get('actual','')})"
                        )
                row.append(f"{r['fields_correct']}/{r['fields_total']}")
                row.append("YES" if r["all_fields_correct"] else "NO")
                writer.writerow(row)

        print(f"\n  Saved: {json_path}")
        print(f"  Saved: {csv_path}")

        sheet_total = len(all_rows)
        sheet_ok = sum(1 for r in all_rows if r["all_fields_correct"])
        sheet_fc = sum(r["fields_correct"] for r in all_rows)
        sheet_ft = sum(r["fields_total"] for r in all_rows)
        print(f"  Sheet overall: {sheet_ok}/{sheet_total} = {sheet_ok/sheet_total*100:.1f}%")
        print(f"  Sheet field:   {sheet_fc}/{sheet_ft} = {sheet_fc/sheet_ft*100:.1f}%")

        if sheet_name == "IV continuous (16)":
            tit_rows = [r for r in all_rows if r.get("entry_type") == "titratable"]
            nontit_rows = [r for r in all_rows if r.get("entry_type") == "non-titratable"]
            if tit_rows:
                tok = sum(1 for r in tit_rows if r["all_fields_correct"])
                print(f"  Titratable:     {tok}/{len(tit_rows)} = {tok/len(tit_rows)*100:.1f}%")
            if nontit_rows:
                ntok = sum(1 for r in nontit_rows if r["all_fields_correct"])
                print(f"  Non-titratable: {ntok}/{len(nontit_rows)} = {ntok/len(nontit_rows)*100:.1f}%")

    # ── Grand summary ──
    g = grand_total
    print(f"\n{'='*72}")
    print(f"GRAND SUMMARY (CoT pipeline, {MODEL_NAME} local, {NUM_RUNS} runs)")
    print(f"  Overall accuracy: {g['correct']}/{g['entries']} = "
          f"{g['correct']/g['entries']*100:.1f}%")
    print(f"  Field accuracy:   {g['fields_ok']}/{g['fields']} = "
          f"{g['fields_ok']/g['fields']*100:.1f}%")
    if tit_stats["entries"]:
        t = tit_stats
        print(f"  IV CI titratable:     {t['correct']}/{t['entries']} = "
              f"{t['correct']/t['entries']*100:.1f}%")
    if nontit_stats["entries"]:
        n = nontit_stats
        print(f"  IV CI non-titratable: {n['correct']}/{n['entries']} = "
              f"{n['correct']/n['entries']*100:.1f}%")
    print(f"{'='*72}")


if __name__ == "__main__":
    main()
