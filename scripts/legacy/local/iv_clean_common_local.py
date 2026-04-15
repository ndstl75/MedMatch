#!/usr/bin/env python3
"""Shared helpers for clean IV-only MedMatch experiments."""

import json
import os
import re
from typing import Dict, Iterable, List, Optional

import openpyxl

from local_llm import OLLAMA_MODEL, chat_completion, resolve_project_file


MODEL_NAME = os.environ.get("OLLAMA_MODEL", OLLAMA_MODEL)
TEMPERATURE = float(os.environ.get("MEDMATCH_TEMPERATURE", "0.1"))
MAX_RETRIES = int(os.environ.get("MEDMATCH_MAX_RETRIES", "3"))
RETRY_DELAY = float(os.environ.get("MEDMATCH_RETRY_DELAY", "5"))
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "1"))
DEFAULT_SHEETS = ["IV push (17)", "IV continuous (16)"]

SHEET_CONFIG = {
    "IV push (17)": {
        "instruction": "IV push",
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of volume": 7,
            "volume unit of measure": 8,
            "concentration of solution": 9,
            "concentration unit of measure": 10,
            "formulation": 11,
            "frequency": 12,
        },
    },
    "IV continuous (16)": {
        "instruction": "IV continuous",
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "starting rate": 10,
            "unit of measure": 11,
            "titration dose": 12,
            "titration unit of measure": 13,
            "titration frequency": 14,
            "titration goal based on physiologic response, laboratory result, or assessment score": 15,
        },
    },
}

EXPECTED_KEYS_BY_SHEET = {
    sheet_name: list(config["ground_truth_cols"].keys())
    for sheet_name, config in SHEET_CONFIG.items()
}


IV_PUSH_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV push dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][amount of volume][volume unit of measure] of the [concentration of solution][concentration unit of measure][formulation] intravenous push [frequency]

[drug name]: The generic or brand name of the medication administered by IV push.
[numerical dose]: The numeric value of the drug dose delivered per administration (e.g., 2, 10).
[abbreviated unit strength of dose]: The standardized abbreviated unit for the dose (e.g., mg, mcg).
[amount of volume]: The numeric volume administered with the IV push (e.g., 2, 5).
[volume unit of measure]: The standardized abbreviated unit for volume (e.g., mL).
[concentration of solution]: The strength of the drug within the solution, normalized to per 1 mL (e.g., 10 if the source says 20 mg/2 mL).
[concentration unit of measure]: The unit basis used to express the concentration (e.g., mg/mL).
[formulation]: The injectable dosage form (e.g., solution).
intravenous push: The fixed route and method of administration.
[frequency]: How often the IV push dose is administered (e.g., every 6 hours, once).

Example of input:
A total of 6mg of adenosine (2 ml) of the 3 mg/ml vial solution intravenous was pushed once
Example of MedMatch JSON format:
{ "drug name": "adenosine",
"numerical dose": 6,
"abbreviated unit strength of dose": "mg",
"amount of volume": 2,
"volume unit of measure": "mL",
"concentration of solution": 3,
"concentration unit of measure": "mg/mL",
"formulation": "vial solution",
"frequency": "once"}
"""


IV_CONTINUOUS_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV continuous dosage form medications is:

Titratable IV CI:
[drug name][numerical dose][abbreviated unit strength of dose] "in" [diluent volume][volume unit of measure][compatible diluent type] "continuous intravenous infusion starting at" [starting rate][unit of measure] "titrated by" [titration dose][titration unit of measure] [titration frequency] to achieve a goal [titration goal based on physiologic response, laboratory result, or assessment score]

Non-titratable IV CI:
[drug name][numerical dose][abbreviated unit strength of dose][diluent volume][volume unit of measure] "in" [compatible diluent type] "continuous intravenous infusion at" [rate][unit of measure]

Definitions:
[drug name]: The generic or brand name of the medication administered as a continuous IV infusion.
[numerical dose]: The total amount of drug contained in the prepared infusion bag or bottle, not the per-mL concentration. For example, if the order says 200 mg/100 mL, write 200 here. If the order says 1 mg/mL in 100 mL, write 100 here.
[abbreviated unit strength of dose]: The standardized abbreviated unit associated with the dose (e.g., mg, units).
[diluent volume]: The total prepared volume of the infusion bag or bottle (e.g., 100, 250). Do not write 0 when the volume is implied by a concentration such as 1 mg/mL in 100 mL.
[volume unit of measure]: The standardized abbreviated unit for the diluent volume (e.g., mL).
[compatible diluent type]: The IV fluid used to dilute the medication (e.g., 0.9% sodium chloride, D5W).
continuous intravenous infusion: The fixed route and method of administration.
[starting rate]: The initial infusion rate at which the medication is started, or the fixed infusion rate if the infusion is non-titratable (e.g., 0.05, 5).
[unit of measure]: The unit associated with the infusion rate (e.g., mcg/kg/min, units/hr, mL/hr).
[titration dose]: The numeric amount by which the infusion rate is adjusted per titration step (e.g., 0.01, 2).
[titration unit of measure]: The unit associated with the titration increment (e.g., mcg/kg/min, units/hr).
[titration frequency]: The time interval between allowable titrations (e.g., every 5 minutes, every 15 minutes).
[titration goal]: Copy the clinical target guiding titration as closely as possible from the sentence instead of paraphrasing (e.g., MAP >= 65 mmHg, RASS score -1 to 1).

Example of titratable input:
The patient was started on Ketamine continuous intravenous infusion at 0.2 mg/kg/hour using a bag of 500 mg/500 ml in 0.9% sodium chloride and titrate by 0.1 mg/kg/hour every 20 minutes to achieve a goal RASS of -4 to -5.
Example of MedMatch JSON Format:
{ "drug name": "ketamine",
"numerical dose": 500,
"abbreviated unit strength of dose": "mg",
"diluent volume": 500,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.2,
"unit of measure": "mg/kg/hr",
"titration dose": 0.1,
"titration unit of measure": "mg/kg/hr",
"titration frequency": "every 20 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5"}

Example of non-titratable input:
The patient was started on a vasopressin continuous intravenous infusion at 0.04 units/minute using a 40 units/100 ml bag in 0.9% sodium chloride.
Example of MedMatch JSON Format:
{ "drug name": "vasopressin",
"numerical dose": 40,
"abbreviated unit strength of dose": "units",
"diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.04,
"unit of measure": "units/minute"}

Example of total-dose interpretation:
The patient received a continuous intravenous infusion of midazolam at a starting rate of 0.5 mg/hr (1 mg/mL in 100 mL of 0.9% sodium chloride), and the infusion was titrated by 0.5 mg/hr every 5 minutes to achieve a RASS goal of -4 to -5.
Example of MedMatch JSON Format:
{ "drug name": "midazolam",
"numerical dose": 100,
"abbreviated unit strength of dose": "mg",
"diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.5,
"unit of measure": "mg/hr",
"titration dose": 0.5,
"titration unit of measure": "mg/hr",
"titration frequency": "every 5 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5"}
"""


IV_EXEMPLAR_BANK = {
    "IV push (17)": [
        {
            "label": "hydralazine",
            "prompt": "The patient recieved hydralazine 10 mg (0.5 mL of the 20 mg/mL vial) and was administered once by intravenous push.",
            "output": {
                "drug name": "Hydralazine",
                "numerical dose": 10,
                "abbreviated unit strength of dose": "mg",
                "amount of volume": 0.5,
                "volume unit of measure": "mL",
                "concentration of solution": 20,
                "concentration unit of measure": "mg/mL",
                "formulation": "vial solution",
                "frequency": "once",
            },
        },
        {
            "label": "dexamethasone",
            "prompt": "Dexamethasone 6 mg, equivalent to 0.6 mL from a 10 mg/mL vial solution, was pushed intravenously daily.",
            "output": {
                "drug name": "Dexamethasone",
                "numerical dose": 6,
                "abbreviated unit strength of dose": "mg",
                "amount of volume": 0.6,
                "volume unit of measure": "mL",
                "concentration of solution": 10,
                "concentration unit of measure": "mg/mL",
                "formulation": "vial solution",
                "frequency": "once daily",
            },
        },
        {
            "label": "famotidine",
            "prompt": "Famotidine 20 mg, 2 mL of a 20 mg/2 mL vial, was administered twice daily via intravenous push.",
            "output": {
                "drug name": "Famotidine",
                "numerical dose": 20,
                "abbreviated unit strength of dose": "mg",
                "amount of volume": 2,
                "volume unit of measure": "mL",
                "concentration of solution": 10,
                "concentration unit of measure": "mg/mL",
                "formulation": "vial solution",
                "frequency": "twice daily",
            },
        },
        {
            "label": "ondansetron",
            "prompt": "Ondansetron 4 mg every 6 hours, prepared from 2 mL of a 2 mg/mL vial solution and was pushed intravenously.",
            "output": {
                "drug name": "Ondansetron",
                "numerical dose": 4,
                "abbreviated unit strength of dose": "mg",
                "amount of volume": 2,
                "volume unit of measure": "mL",
                "concentration of solution": 2,
                "concentration unit of measure": "mg/mL",
                "formulation": "vial solution",
                "frequency": "every 6 hours",
            },
        },
        {
            "label": "fentanyl",
            "prompt": "Fentanyl 50 mcg was administered by intravenous push every 4 hours, using 1 mL from the 50 mcg/mL vial solution.",
            "output": {
                "drug name": "Fentanyl",
                "numerical dose": 50,
                "abbreviated unit strength of dose": "mcg",
                "amount of volume": 1,
                "volume unit of measure": "mL",
                "concentration of solution": 50,
                "concentration unit of measure": "mcg/mL",
                "formulation": "vial solution",
                "frequency": "every 4 hours",
            },
        },
    ],
    "IV continuous (16)": [
        {
            "label": "dexmedetomidine",
            "prompt": "Dexmedetomidine, diluted to 1000 mcg in 250 mL 0.9% sodium chloride, was used to achieve a goal RASS of -1 to -2; the continuous intravenous infusion was started at 0.2 mcg/kg/hr and titrated by 0.2 mcg/kg/hr every 20 minute.",
            "output": {
                "drug name": "Dexmedetomidine",
                "numerical dose": 1000,
                "abbreviated unit strength of dose": "mcg",
                "diluent volume": 250,
                "volume unit of measure": "mL",
                "compatible diluent type": "0.9% sodium chloride",
                "starting rate": 0.2,
                "unit of measure": "mcg/kg/hr",
                "titration dose": 0.2,
                "titration unit of measure": "mcg/kg/hr",
                "titration frequency": "every 20 minutes",
                "titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -1 to -2",
            },
        },
        {
            "label": "lorazepam",
            "prompt": "A lorazepam infusion, prepared as 1 mg/mL in D5W, was started as a continuous intravenous infusion at 0.5 mg/hr and titrated by 0.5 mg/hr every 10 minutes to achieve a goal RASS of –4 to –5.",
            "output": {
                "drug name": "Lorazepam",
                "numerical dose": 1,
                "abbreviated unit strength of dose": "mg",
                "diluent volume": 1,
                "volume unit of measure": "mL",
                "compatible diluent type": "D5W",
                "starting rate": 0.5,
                "unit of measure": "mg/hr",
                "titration dose": 0.5,
                "titration unit of measure": "mg/hr",
                "titration frequency": "every 10 minutes",
                "titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5",
            },
        },
        {
            "label": "phenylephrine",
            "prompt": "Phenylephrine 100 mg in 250 mL of 0.9% sodium chloride was administered as a continuous intravenous infusion, starting at 0.2 mcg/kg/min and titrated by 0.1 mcg/kg/min every minute to achieve a goal MAP of 65 mmHg or greater.",
            "output": {
                "drug name": "Phenylephrine",
                "numerical dose": 100,
                "abbreviated unit strength of dose": "mg",
                "diluent volume": 250,
                "volume unit of measure": "mL",
                "compatible diluent type": "0.9% sodium chloride",
                "starting rate": 0.2,
                "unit of measure": "mcg/kg/min",
                "titration dose": 0.1,
                "titration unit of measure": "mcg/kg/min",
                "titration frequency": "every 1 minute",
                "titration goal based on physiologic response, laboratory result, or assessment score": "MAP of 65 mmHg or greater",
            },
        },
        {
            "label": "norepinephrine",
            "prompt": "Norepinephrine, mixed as 16 mg in 250 mL of 0.9% sodium chloride, was started at 0.05 mcg/kg/min continuous intravenous infusion and titrated by 0.01 mcg/kg/min every 5 minutes to achieve a MAP of 65 mmHg or greater.",
            "output": {
                "drug name": "Norepinephrine",
                "numerical dose": 15,
                "abbreviated unit strength of dose": "mg",
                "diluent volume": 250,
                "volume unit of measure": "mL",
                "compatible diluent type": "0.9% sodium chloride",
                "starting rate": 0.05,
                "unit of measure": "mcg/kg/min",
                "titration dose": 0.01,
                "titration unit of measure": "mcg/kg/min",
                "titration frequency": "every 5 minutes",
                "titration goal based on physiologic response, laboratory result, or assessment score": "MAP of 65 mmHg or greater",
            },
        },
        {
            "label": "milrinone",
            "prompt": "Milrinone was given as a fixed continuous intravenous infusion at 0.125 mcg/kg/min from a 20 mg/100 mL D5W formulation.",
            "output": {
                "drug name": "Milrinone",
                "numerical dose": 20,
                "abbreviated unit strength of dose": "mg",
                "diluent volume": 100,
                "volume unit of measure": "mL",
                "compatible diluent type": "D5W",
                "starting rate": 0.125,
                "unit of measure": "mcg/kg/min",
                "titration dose": "",
                "titration unit of measure": "",
                "titration frequency": "",
                "titration goal based on physiologic response, laboratory result, or assessment score": "",
            },
        },
        {
            "label": "hydromorphone",
            "prompt": "Hydromorphone continuous intravenous infusion therapy, prepared as 50 mg/50 mL 0.9% sodium chloride, was started at 0.2 mg/hr and titrated by 0.2 mg/hr adjustments every 10 minutes to achieve a goal RASS score of -1 to -2.",
            "output": {
                "drug name": "Hydromorphone",
                "numerical dose": 50,
                "abbreviated unit strength of dose": "mg",
                "diluent volume": 50,
                "volume unit of measure": "mL",
                "compatible diluent type": "0.9% sodium chloride",
                "starting rate": 0.2,
                "unit of measure": "mg/hr",
                "titration dose": 0.2,
                "titration unit of measure": "mg/hr",
                "titration frequency": "every 10 minutes",
                "titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -1 to -2",
            },
        },
    ],
}


def selected_sheets_from_env(default: Optional[List[str]] = None) -> List[str]:
    raw = os.environ.get("MEDMATCH_SHEETS", "")
    selected = [s.strip() for s in raw.split(",") if s.strip()]
    return selected or list(default or DEFAULT_SHEETS)


def load_dataset(
    xlsx_path: str,
    selected_sheets: Optional[Iterable[str]] = None,
    max_entries_per_sheet: int = 0,
) -> Dict[str, List[dict]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    selected = set(selected_sheets or DEFAULT_SHEETS)
    dataset: Dict[str, List[dict]] = {}

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in selected:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[config["prompt_col"] - 1].value
            if not prompt:
                break
            ground_truth = {
                key: (row[col_idx - 1].value if row[col_idx - 1].value is not None else "")
                for key, col_idx in config["ground_truth_cols"].items()
            }
            rows.append(
                {
                    "medication": row[0].value,
                    "prompt": str(prompt).strip(),
                    "ground_truth": ground_truth,
                }
            )
            if max_entries_per_sheet and len(rows) >= max_entries_per_sheet:
                break
        dataset[sheet_name] = rows
        print(f"Loaded {len(rows)} rows from {sheet_name}")
    return dataset


def build_response_schema(expected_keys: List[str]) -> dict:
    return {
        "type": "object",
        "properties": {k: {"type": ["string", "number", "integer", "null"]} for k in expected_keys},
        "required": expected_keys,
        "additionalProperties": False,
    }


def parse_json_response(text: str):
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3].strip()
        elif "```" in text:
            text = text[: text.rfind("```")].strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            start = text.index("{")
            end = text.rindex("}") + 1
            return json.loads(text[start:end])
        except (ValueError, json.JSONDecodeError):
            return None


def coerce_output_object(parsed, expected_keys: List[str]) -> dict:
    if parsed is None:
        return {key: "" for key in expected_keys}
    if isinstance(parsed, list):
        if len(parsed) == 1 and isinstance(parsed[0], dict):
            parsed = parsed[0]
        else:
            parsed = dict(zip(expected_keys, parsed))
    if not isinstance(parsed, dict):
        return {key: "" for key in expected_keys}
    return {key: parsed.get(key, "") for key in expected_keys}


def exact_value(value) -> str:
    return "" if value is None else str(value)


def normalize(value) -> str:
    """Strict: lowercase + whitespace-collapse only. No unit rewrites, no
    unicode substitution, no numeric coercion. Local-side rewrites beyond
    case/whitespace would relax the string-exact match standard."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def compare_results_exact(llm_output, ground_truth: dict) -> dict:
    results = {}
    for key, expected in ground_truth.items():
        actual = llm_output.get(key, "") if llm_output else ""
        results[key] = {
            "expected": normalize(expected),
            "actual": normalize(actual),
            "match": normalize(expected) == normalize(actual),
        }
    return results


def all_fields_match(comparison: dict) -> bool:
    return all(v["match"] for v in comparison.values())


def call_local_model(system_prompt: str, user_prompt: str, *, model: str = None, temperature: float = None, retries: int = None):
    model_name = model or MODEL_NAME
    temp = TEMPERATURE if temperature is None else temperature
    retry_count = MAX_RETRIES if retries is None else retries

    for attempt in range(retry_count):
        try:
            return chat_completion(
                system_prompt,
                user_prompt,
                temperature=temp,
                model=model_name,
            )
        except Exception as exc:
            if attempt < retry_count - 1:
                print(f"    retry {attempt + 1}: {exc}")
                import time

                time.sleep(RETRY_DELAY)
            else:
                raise


def render_examples(examples: List[dict]) -> str:
    parts = []
    for idx, ex in enumerate(examples, 1):
        parts.append(
            f"Example {idx} input:\n{ex['prompt']}\n"
            f"Example {idx} output:\n{json.dumps(ex['output'], ensure_ascii=False)}"
        )
    return "\n\n".join(parts)


def tokenize_for_retrieval(text: str) -> set:
    return set(re.findall(r"[a-z0-9%/]+", text.lower()))


def retrieve_topk_examples(query: str, bank: List[dict], k: int = 2, exclude_labels: Optional[Iterable[str]] = None) -> List[dict]:
    exclude = {label.lower() for label in (exclude_labels or [])}
    q_tokens = tokenize_for_retrieval(query)
    scored = []
    for ex in bank:
        if ex["label"].lower() in exclude:
            continue
        ex_tokens = tokenize_for_retrieval(ex["prompt"])
        union = q_tokens | ex_tokens
        score = len(q_tokens & ex_tokens) / len(union) if union else 0.0
        scored.append((score, ex))
    scored.sort(key=lambda item: (-item[0], item[1]["label"]))
    return [ex for _, ex in scored[:k]]
