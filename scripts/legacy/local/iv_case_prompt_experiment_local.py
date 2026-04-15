#!/usr/bin/env python3
"""
Targeted IV experiment for local Gemma/Ollama.

Focus:
- Keep the original MedMatch task and scoring
- Add IV-specific prompt constraints and a few targeted examples
- Apply only conservative post-normalization for wording/unit drift

This is meant to test the user's hypothesis that a small amount of
case-guidance + prompt tightening can improve IV extraction.
"""

import csv
import json
import os
import re
import time
from datetime import datetime

import openpyxl

from local_llm import OLLAMA_MODEL, chat_completion, resolve_project_file


MODEL_NAME = os.environ.get("OLLAMA_MODEL", OLLAMA_MODEL)
TEMPERATURE = 0.1
NUM_RUNS = int(os.environ.get("MEDMATCH_NUM_RUNS", "3"))  # triplicate per paper
MAX_RETRIES = 3
RETRY_DELAY = float(os.environ.get("MEDMATCH_RETRY_DELAY", "5"))
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "0.5"))
SELECTED_SHEETS = [
    s.strip() for s in os.environ.get("MEDMATCH_SHEETS", "").split(",") if s.strip()
]
MAX_ENTRIES_PER_SHEET = int(os.environ.get("MEDMATCH_MAX_ENTRIES", "0"))

SYSTEM_PROMPT = (
    "You are a clinical pharmacist who formats medication orders. "
    "Only output the MedMatch JSON format."
)

IV_INTERMITTENT_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV intermittent dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][amount of diluent volume][volume unit of measure][compatible diluent type] intravenously infused over [infusion time] [frequency]

[drug name]: The generic or brand name of the medication to be administered intravenously.
[numerical dose]: The numeric value of the drug dose to be given per administration (e.g., 1, 500).
[abbreviated unit strength of dose]: The standardized abbreviated unit associated with the dose (e.g., mg, g, units).
[amount of diluent volume]: The numeric volume of diluent used to prepare the IV medication (e.g., 50, 100).
[volume unit of measure]: The standardized abbreviated unit for diluent volume (e.g., mL).
[compatible diluent type]: The IV fluid used for dilution that is compatible with the medication (e.g., 0.9% sodium chloride, D5W).
intravenous: The fixed route of administration.
infused over: Indicates the medication is administered as an infusion rather than IV push.
[infusion time]: The duration over which the medication is infused (e.g., 30 minutes, 1 hour).
[frequency]: How often the intermittent IV dose is administered (e.g., every 8 hours, once daily).

Example of input:
Cefepime 2000 mg was delivered as a 30 minute intravenous infusion, prepared in 100 mL of 0.9% sodium chloride and administered every 8 hours.
Example of MedMatch JSON Format:
{ "drug name": "cefepime",
"numerical dose": 2000,
"abbreviated unit strength of dose": "mg",
"amount of diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"infusion time": "30 minutes",
"frequency": "every 8 hours"}
"""

IV_PUSH_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV push dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][amount of volume][volume unit of measure] of the [concentration of solution][concentration unit of measure][formulation] intravenous push [frequency]

[drug name]: The generic or brand name of the medication administered by IV push.
[numerical dose]: The numeric value of the drug dose delivered per administration (e.g., 2, 10).
[abbreviated unit strength of dose]: The standardized abbreviated unit for the dose (e.g., mg, mcg).
[amount of volume]: The numeric volume administered with the IV push (e.g., 2, 5).
[volume unit of measure]: The standardized abbreviated unit for volume (e.g., mL).
[concentration of solution]: The strength of the drug within the solution, normalized to per 1 mL. For example, if the sentence says 20 mg/2 mL, write 10 here.
[concentration unit of measure]: The unit basis used to express the concentration (e.g., mg/mL, mcg/mL).
[formulation]: The injectable dosage form. If the source describes a vial injection, use "vial solution".
intravenous push: The fixed route and method of administration.
[frequency]: How often the IV push dose is administered (e.g., every 6 hours, once, once daily).

Example of input:
A total of 6mg of adenosine (2 ml) of the 3 mg/ml vial solution intravenous was pushed once
Example of MedMatch JSON format:
{ "drug name": "adenosine",
"numerical dose": 6,
"abbreviated unit strength of dose": "mg",
"amount of volume": 2,
"volume unit of measure": "mL",
"concentration of solution": 3,
"concentration unit of measure": "mg/mL",
"formulation": "vial solution",
"frequency": "once"}

Example of concentration conversion:
Famotidine 20 mg, 2 mL of a 20 mg/2 mL vial, was administered twice daily via intravenous push.
Example of MedMatch JSON format:
{ "drug name": "famotidine",
"numerical dose": 20,
"abbreviated unit strength of dose": "mg",
"amount of volume": 2,
"volume unit of measure": "mL",
"concentration of solution": 10,
"concentration unit of measure": "mg/mL",
"formulation": "vial solution",
"frequency": "twice daily"}

Example of once-daily wording:
Dexamethasone 6 mg, equivalent to 0.6 mL from a 10 mg/mL vial solution, was pushed intravenously daily.
Example of MedMatch JSON format:
{ "drug name": "dexamethasone",
"numerical dose": 6,
"abbreviated unit strength of dose": "mg",
"amount of volume": 0.6,
"volume unit of measure": "mL",
"concentration of solution": 10,
"concentration unit of measure": "mg/mL",
"formulation": "vial solution",
"frequency": "once daily"}
"""

IV_CONTINUOUS_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV continuous dosage form medications is:

Titratable IV CI:
[drug name][numerical dose][abbreviated unit strength of dose] "in" [diluent volume][volume unit of measure][compatible diluent type] "continuous intravenous infusion starting at" [starting rate][unit of measure] "titrated by" [titration dose][titration unit of measure] [titration frequency] to achieve a goal [titration goal based on physiologic response, laboratory result, or assessment score]

Non-titratable IV CI:
[drug name][numerical dose][abbreviated unit strength of dose][diluent volume][volume unit of measure] "in" [compatible diluent type] "continuous intravenous infusion at" [rate][unit of measure]

Definitions:
[drug name]: The generic or brand name of the medication administered as a continuous IV infusion.
[numerical dose]: If the sentence gives a prepared bag such as 4 mg/250 mL, use the total amount in the bag (4). If the sentence gives only a concentration with no total bag volume, copy the numeric concentration value instead of inventing a total bag amount.
[abbreviated unit strength of dose]: For prepared bags, use the total-dose unit (e.g., mg, units). If the sentence only gives a concentration with no total bag volume and the dataset expects the concentration wording, preserve that concentration unit wording exactly.
[diluent volume]: The total prepared volume of the infusion bag or bottle (e.g., 100, 250). If the sentence does not provide a total volume, leave it empty unless the formulation is explicitly written as 1 mg/mL style and you are preserving a 1 mL concentration representation.
[volume unit of measure]: The standardized abbreviated unit for the diluent volume (e.g., mL).
[compatible diluent type]: The IV fluid used to dilute the medication (e.g., 0.9% sodium chloride, D5W).
continuous intravenous infusion: The fixed route and method of administration.
[starting rate]: The initial infusion rate at which the medication is started, or the fixed infusion rate if the infusion is non-titratable.
[unit of measure]: The unit associated with the infusion rate (e.g., mcg/kg/min, units/hr, mg/hr).
[titration dose]: The numeric amount by which the infusion rate is adjusted per titration step.
[titration unit of measure]: The unit associated with the titration increment.
[titration frequency]: Copy the explicit time interval in normalized form such as "every 1 minute" or "every 5 minutes".
[titration goal]: Copy the clinical target as tightly as possible. Do not add extra words like "goal", "score", or "light sedation with" if they are not needed in the final value.

Example of titratable input:
The patient was started on Ketamine continuous intravenous infusion at 0.2 mg/kg/hour using a bag of 500 mg/500 ml in 0.9% sodium chloride and titrate by 0.1 mg/kg/hour every 20 minutes to achieve a goal RASS of -4 to -5.
Example of MedMatch JSON Format:
{ "drug name": "ketamine",
"numerical dose": 500,
"abbreviated unit strength of dose": "mg",
"diluent volume": 500,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.2,
"unit of measure": "mg/kg/hr",
"titration dose": 0.1,
"titration unit of measure": "mg/kg/hr",
"titration frequency": "every 20 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5"}

Example of concentration-only titratable input:
A propofol continuous intravenous infusion, from a 10 mg/mL concentration, was started at 5 mcg/kg/min and increased by 5 mcg/kg/min steps every 5 minutes to maintain light sedation with a RASS of -1 to -2.
Example of MedMatch JSON Format:
{ "drug name": "propofol",
"numerical dose": 10,
"abbreviated unit strength of dose": "mg/mL",
"diluent volume": "",
"volume unit of measure": "",
"compatible diluent type": "",
"starting rate": 5,
"unit of measure": "mcg/kg/min",
"titration dose": 5,
"titration unit of measure": "mcg/kg/min",
"titration frequency": "every 5 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -1 to -2"}

Example of concentration-only with diluent type:
A lorazepam infusion, prepared as 1 mg/mL in D5W, was started as a continuous intravenous infusion at 0.5 mg/hr and titrated by 0.5 mg/hr every 10 minutes to achieve a goal RASS of -4 to -5.
Example of MedMatch JSON Format:
{ "drug name": "lorazepam",
"numerical dose": 1,
"abbreviated unit strength of dose": "mg",
"diluent volume": 1,
"volume unit of measure": "mL",
"compatible diluent type": "D5W",
"starting rate": 0.5,
"unit of measure": "mg/hr",
"titration dose": 0.5,
"titration unit of measure": "mg/hr",
"titration frequency": "every 10 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5"}

Example of total-dose interpretation:
The patient received a continuous intravenous infusion of midazolam at a starting rate of 0.5 mg/hr (1 mg/mL in 100 mL of 0.9% sodium chloride), and the infusion was titrated by 0.5 mg/hr every 5 minutes to achieve a RASS goal of -4 to -5.
Example of MedMatch JSON Format:
{ "drug name": "midazolam",
"numerical dose": 100,
"abbreviated unit strength of dose": "mg",
"diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.5,
"unit of measure": "mg/hr",
"titration dose": 0.5,
"titration unit of measure": "mg/hr",
"titration frequency": "every 5 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5"}

Example of non-titratable input:
The patient was started on a vasopressin continuous intravenous infusion at 0.04 units/minute using a 40 units/100 ml bag in 0.9% sodium chloride.
Example of MedMatch JSON Format:
{ "drug name": "vasopressin",
"numerical dose": 40,
"abbreviated unit strength of dose": "units",
"diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.04,
"unit of measure": "units/minute",
"titration dose": "",
"titration unit of measure": "",
"titration frequency": "",
"titration goal based on physiologic response, laboratory result, or assessment score": ""}
"""

SHEET_CONFIG = {
    "IV intermittent (16)": {
        "instruction": IV_INTERMITTENT_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "infusion time": 10,
            "frequency": 11,
        },
    },
    "IV push (17)": {
        "instruction": IV_PUSH_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of volume": 7,
            "volume unit of measure": 8,
            "concentration of solution": 9,
            "concentration unit of measure": 10,
            "formulation": 11,
            "frequency": 12,
        },
    },
    "IV continuous (16)": {
        "instruction": IV_CONTINUOUS_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "starting rate": 10,
            "unit of measure": 11,
            "titration dose": 12,
            "titration unit of measure": 13,
            "titration frequency": 14,
            "titration goal based on physiologic response, laboratory result, or assessment score": 15,
        },
    },
}


def normalize(value):
    """Strict: lowercase + whitespace-collapse only. No unit rewrites, no
    unicode substitution, no numeric coercion. Local-side rewrites beyond
    case/whitespace would relax the string-exact match standard."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def load_dataset(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    for sheet_name, cfg in SHEET_CONFIG.items():
        if SELECTED_SHEETS and sheet_name not in SELECTED_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[cfg["prompt_col"] - 1].value
            if not prompt:
                break
            gt = {}
            for key, col in cfg["ground_truth_cols"].items():
                val = row[col - 1].value
                gt[key] = val if val is not None else ""
            rows.append(
                {
                    "medication": row[0].value,
                    "prompt": str(prompt).strip(),
                    "ground_truth": gt,
                }
            )
            if MAX_ENTRIES_PER_SHEET and len(rows) >= MAX_ENTRIES_PER_SHEET:
                break
        dataset[sheet_name] = rows
        print(f"Loaded {len(rows)} rows from {sheet_name}")
    return dataset


def build_schema(expected_keys):
    return {
        "type": "object",
        "properties": {k: {"type": ["string", "number", "integer", "null"]} for k in expected_keys},
        "required": expected_keys,
        "additionalProperties": False,
    }


def build_instruction(sheet_name, base_instruction, prompt, expected_keys):
    lines = [
        "Return one JSON object only.",
        "Do not wrap the JSON in markdown.",
        f"Use exactly these keys in this order: {', '.join(expected_keys)}.",
    ]

    if sheet_name == "IV intermittent (16)":
        lines.extend(
            [
                'If frequency means daily, write "once daily".',
                "Preserve full as-needed qualifiers such as 'as needed for phosphorous replacement'.",
            ]
        )
    elif sheet_name == "IV push (17)":
        lines.extend(
            [
                'If the source says vial or vial solution, output formulation as "vial solution".',
                "Normalize concentration to per 1 mL. Example: 20 mg/2 mL -> concentration of solution 10 and concentration unit of measure mg/mL.",
                'If frequency means daily, write "once daily".',
            ]
        )
    elif sheet_name == "IV continuous (16)":
        prompt_lower = prompt.lower()
        if any(token in prompt_lower for token in ("titrate", "titrated", "goal", "increase")):
            lines.append("This is a titratable continuous infusion. Use the key 'starting rate' for the initial rate.")
        else:
            lines.append("This is a non-titratable continuous infusion. Leave all titration fields as empty strings.")
        lines.extend(
            [
                "When the target says every minute, normalize it as every 1 minute.",
                "Copy titration goals tightly and avoid adding filler words like goal, score, or light sedation with.",
            ]
        )

    return f"{base_instruction}\n\n" + "\n".join(lines)


def call_model(sheet_name, instruction, prompt, expected_keys):
    prompt_full = (
        f"{build_instruction(sheet_name, instruction, prompt, expected_keys)}\n\n"
        f"Now process this medication order:\n{prompt}"
    )
    schema_hint = json.dumps(build_schema(expected_keys), ensure_ascii=False)
    final_prompt = (
        f"{prompt_full}\n\n"
        "Return valid JSON only.\n"
        "Do not wrap the JSON in markdown.\n"
        f"Follow this JSON schema exactly:\n{schema_hint}"
    )

    for attempt in range(MAX_RETRIES):
        text = ""
        try:
            text = chat_completion(
                SYSTEM_PROMPT,
                final_prompt,
                temperature=TEMPERATURE,
                model=MODEL_NAME,
            )
            if text.startswith("```"):
                text = text.split("\n", 1)[1] if "\n" in text else text[3:]
                if text.endswith("```"):
                    text = text[:-3].strip()
                elif "```" in text:
                    text = text[:text.rfind("```")].strip()
            obj = json.loads(text)
            if not isinstance(obj, dict):
                obj = {}
            return {k: obj.get(k, "") for k in expected_keys}, text
        except Exception:
            try:
                start = text.index("{")
                end = text.rindex("}") + 1
                obj = json.loads(text[start:end])
                return {k: obj.get(k, "") for k in expected_keys}, text
            except Exception:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                else:
                    return None, text


# NOTE: The previous version had parse_strength_per_ml(), parse_simple_concentration(),
# normalize_goal_text(), and normalize_iv_output() — a post-processor that rewrote the
# LLM's output before scoring (e.g., "daily" -> "once daily", "vial" -> "vial solution",
# "every minute" -> "every 1 minute", titration goals regex-stripped, concentration
# values parsed out of the prompt and substituted into the output).
#
# This is scorer-side post-processing that fixes the model's answers before comparing
# to ground truth. It is the strongest form of grading-standard relaxation. Removed
# entirely so the scorer compares raw model output to GT.


def compare_results(llm_output, ground_truth):
    results = {}
    for key, expected in ground_truth.items():
        e = normalize(expected)
        a = normalize(llm_output.get(key, "") if llm_output else "")
        results[key] = {"expected": e, "actual": a, "match": e == a}
    return results


def all_fields_match(comp):
    return all(v["match"] for v in comp.values())


def main():
    xlsx = resolve_project_file("MedMatch Dataset for Experiment_ Final.xlsx")
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
    os.makedirs(results_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    dataset = load_dataset(xlsx)
    print(f"Model={MODEL_NAME}, temp={TEMPERATURE}, runs={NUM_RUNS}")

    for sheet_name, entries in dataset.items():
        cfg = SHEET_CONFIG[sheet_name]
        expected_keys = list(cfg["ground_truth_cols"].keys())
        all_rows = []
        total_ok = 0
        total_n = 0
        total_fields_ok = 0
        total_fields_n = 0
        print(f"\n=== {sheet_name} | {len(entries)} rows x {NUM_RUNS} runs ===")
        for run in range(1, NUM_RUNS + 1):
            run_ok = 0
            for idx, entry in enumerate(entries, 1):
                print(f"[{run}/{NUM_RUNS}] [{idx}/{len(entries)}] {entry['medication']}...", end=" ", flush=True)
                raw_obj, raw_text = call_model(sheet_name, cfg["instruction"], entry["prompt"], expected_keys)
                obj = raw_obj  # raw model output, no post-processing before scoring
                comp = compare_results(obj, entry["ground_truth"])
                fields_ok = sum(1 for v in comp.values() if v["match"])
                all_ok = all_fields_match(comp)
                print(f"{fields_ok}/{len(expected_keys)}" + (" [ALL]" if all_ok else ""))

                total_fields_ok += fields_ok
                total_fields_n += len(expected_keys)
                total_n += 1
                if all_ok:
                    total_ok += 1
                    run_ok += 1

                all_rows.append(
                    {
                        "run": run,
                        "medication": entry["medication"],
                        "prompt": entry["prompt"],
                        "ground_truth": entry["ground_truth"],
                        "llm_output": obj,
                        "raw_response": raw_text,
                        "comparison": comp,
                        "fields_correct": fields_ok,
                        "fields_total": len(expected_keys),
                        "all_fields_correct": all_ok,
                    }
                )
                time.sleep(SLEEP_BETWEEN_CALLS)
            print(f"  Run {run}: {run_ok}/{len(entries)} ({(run_ok/len(entries)*100):.1f}%)")

        overall = (total_ok / total_n * 100) if total_n else 0
        f_acc = (total_fields_ok / total_fields_n * 100) if total_fields_n else 0
        safe = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        jpath = os.path.join(results_dir, f"{safe}_iv_case_prompt_{ts}.json")
        cpath = os.path.join(results_dir, f"{safe}_iv_case_prompt_{ts}.csv")
        with open(jpath, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, indent=2, default=str)
        with open(cpath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Run", "Medication"] + expected_keys + ["Fields Correct", "All Correct"])
            for r in all_rows:
                row = [r["run"], r["medication"]]
                for k in expected_keys:
                    c = r["comparison"][k]
                    row.append("MATCH" if c["match"] else f"MISS (exp: {c['expected']}, got: {c['actual']})")
                row += [f"{r['fields_correct']}/{r['fields_total']}", "YES" if r["all_fields_correct"] else "NO"]
                writer.writerow(row)
        print(f"Saved: {jpath}")
        print(f"Saved: {cpath}")
        print(f"Summary: overall={overall:.1f}% | field={f_acc:.1f}%")


if __name__ == "__main__":
    main()
