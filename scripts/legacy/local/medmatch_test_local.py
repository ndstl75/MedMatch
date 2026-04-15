#!/usr/bin/env python3
"""
MedMatch LLM Evaluation Script
Sends medication order sentences to a locally hosted Gemma model via Ollama
and evaluates structured output against ground truth from the dataset.
Per the paper: one-shot prompting, run in triplicate, temperature=0.1.
"""

import os
import json
import time
import csv
import re
from datetime import datetime
import openpyxl

from local_llm import OLLAMA_MODEL, chat_completion, resolve_project_file

# ── Config ────────────────────────────────────────────────────────────────────
MODEL_NAME = os.environ.get("OLLAMA_MODEL", OLLAMA_MODEL)
TEMPERATURE = 0.1
NUM_RUNS = int(os.environ.get("MEDMATCH_NUM_RUNS", "3"))  # triplicate per paper by default
MAX_RETRIES = 3
RETRY_DELAY = float(os.environ.get("MEDMATCH_RETRY_DELAY", "5"))  # seconds
SELECTED_SHEETS = [
    s.strip() for s in os.environ.get("MEDMATCH_SHEETS", "").split(",") if s.strip()
]
MAX_ENTRIES_PER_SHEET = int(os.environ.get("MEDMATCH_MAX_ENTRIES", "0"))
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "1"))

# ── Prompts per route category ────────────────────────────────────────────────

SYSTEM_PROMPT = (
    "You are a clinical pharmacist who formats medication orders. "
    "Only output the MedMatch JSON format."
)

PO_SOLID_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for oral solid dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][amount][formulation] by mouth [frequency]

[drug name]: The generic or brand name of the medication.
[numerical dose]: The numeric value of the strength per unit (e.g., 5, 10, 500).
[abbreviated unit strength of dose]: The standardized abbreviated unit associated with the dose (e.g., mg, mcg, g).
[amount]: The number of dosage units taken per administration (e.g., 1, 2).
[formulation]: The oral solid dosage form (e.g., tablet, capsule, extended-release tablet). Copy the dosage-form wording from the order as closely as possible, including qualifiers such as extended-release or delayed-release.
by mouth: The route of administration, fixed as oral.
[frequency]: How often the medication is taken (e.g., once daily, twice daily, every 8 hours). Preserve the full schedule phrase, including qualifiers such as as needed, at bedtime, or indication text if present.

Example of input:
Administer oral benztropine four times daily as needed, a dose of 1mg (1 tablet).
Example of MedMatch JSON format:
{ "drug name": "benztropine",
"numerical dose": 1,
"abbreviated unit strength of dose": "mg",
"amount": 1,
"formulation": "tablet",
"route": "by mouth",
"frequency": "four times daily as needed"}

Example of preserving full frequency text:
Administer oral methocarbamol, 750 mg (1 tablet), three times daily as needed for muscle spasms.
Example of MedMatch JSON format:
{ "drug name": "methocarbamol",
"numerical dose": 750,
"abbreviated unit strength of dose": "mg",
"amount": 1,
"formulation": "tablet",
"route": "by mouth",
"frequency": "three times daily as needed for muscle spasms"}
"""

PO_LIQUID_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for oral liquid dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][numerical volume][abbreviated unit strength of volume] of the [concentration of formulation][formulation unit of measure][formulation] by mouth [frequency]

[drug name]: The generic or brand name of the medication.
[numerical dose]: The numeric amount of drug administered per dose (e.g., 250, 5).
[abbreviated unit strength of dose]: The standardized abbreviated unit for the drug dose (e.g., mg, mcg).
[numerical volume]: The numeric volume administered per dose (e.g., 5, 10).
[abbreviated unit strength of volume]: The standardized abbreviated unit for volume (e.g., mL).
[concentration of formulation]: The strength of the medication per 1 mL. For example, if the sentence says 250 mg/5 mL, write 50 here.
[formulation unit of measure]: The unit used in the concentration denominator (e.g., mL).
[formulation]: The oral liquid dosage form (e.g., solution, suspension, syrup).
[route]: The route of administration, fixed as oral.
[frequency]: How often the medication is administered (e.g., once daily, every 6 hours).

Example of input:
An oral solution of diazepam (1mg/mL) is used to administer 5mg (5mL) dose by mouth once daily.
Example of MedMatch JSON Format:
{ "drug name": "diazepam",
"numerical dose": 5,
"abbreviated unit strength of dose": "mg",
"numerical volume": 5,
"volume unit of measure": "mL",
"concentration of formulation": 1,
"formulation unit of measure": "mg/mL",
"formulation": "solution",
"route": "by mouth",
"frequency": "once daily"}

Example of concentration conversion:
Using a 250mg/5mL solution, administer a dose of 500mg of valproic acid (10mL) every 6 hours by mouth.
Example of MedMatch JSON Format:
{ "drug name": "valproic acid",
"numerical dose": 500,
"abbreviated unit strength of dose": "mg",
"numerical volume": 10,
"volume unit of measure": "mL",
"concentration of formulation": 50,
"formulation unit of measure": "mg/mL",
"formulation": "solution",
"route": "by mouth",
"frequency": "every 6 hours"}
"""

IV_INTERMITTENT_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV intermittent dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][amount of diluent volume][volume unit of measure][compatible diluent type] intravenously infused over [infusion time] [frequency]

[drug name]: The generic or brand name of the medication to be administered intravenously.
[numerical dose]: The numeric value of the drug dose to be given per administration (e.g., 1, 500).
[abbreviated unit strength of dose]: The standardized abbreviated unit associated with the dose (e.g., mg, g, units).
[amount of diluent volume]: The numeric volume of diluent used to prepare the IV medication (e.g., 50, 100).
[volume unit of measure]: The standardized abbreviated unit for diluent volume (e.g., mL).
[compatible diluent type]: The IV fluid used for dilution that is compatible with the medication (e.g., 0.9% sodium chloride, D5W).
intravenous: The fixed route of administration.
infused over: Indicates the medication is administered as an infusion rather than IV push.
[infusion time]: The duration over which the medication is infused (e.g., 30 minutes, 1 hour).
[frequency]: How often the intermittent IV dose is administered (e.g., every 8 hours, once daily).

Example of input:
Cefepime 2000 mg was delivered as a 30 minute intravenous infusion, prepared in 100 mL of 0.9% sodium chloride and administered every 8 hours.
Example of MedMatch JSON Format:
{ "drug name": "cefepime",
"numerical dose": 2000,
"abbreviated unit strength of dose": "mg",
"amount of diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"infusion time": "30 minutes",
"frequency": "every 8 hours"}
"""

IV_PUSH_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV push dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][amount of volume][volume unit of measure] of the [concentration of solution][concentration unit of measure][formulation] intravenous push [frequency]

[drug name]: The generic or brand name of the medication administered by IV push.
[numerical dose]: The numeric value of the drug dose delivered per administration (e.g., 2, 10).
[abbreviated unit strength of dose]: The standardized abbreviated unit for the dose (e.g., mg, mcg).
[amount of volume]: The numeric volume administered with the IV push (e.g., 2, 5).
[volume unit of measure]: The standardized abbreviated unit for volume (e.g., mL).
[concentration of solution]: The strength of the drug within the solution, normalized to per 1 mL (e.g., 10 if the source says 20 mg/2 mL).
[concentration unit of measure]: The unit basis used to express the concentration (e.g., mg/mL).
[formulation]: The injectable dosage form (e.g., solution).
intravenous push: The fixed route and method of administration.
[frequency]: How often the IV push dose is administered (e.g., every 6 hours, once).

Example of input:
A total of 6mg of adenosine (2 ml) of the 3 mg/ml vial solution intravenous was pushed once
Example of MedMatch JSON format:
{ "drug name": "adenosine",
"numerical dose": 6,
"abbreviated unit strength of dose": "mg",
"amount of volume": 2,
"volume unit of measure": "mL",
"concentration of solution": 3,
"concentration unit of measure": "mg/mL",
"formulation": "vial solution",
"frequency": "once"}
"""

IV_CONTINUOUS_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for IV continuous dosage form medications is:

Titratable IV CI:
[drug name][numerical dose][abbreviated unit strength of dose] "in" [diluent volume][volume unit of measure][compatible diluent type] "continuous intravenous infusion starting at" [starting rate][unit of measure] "titrated by" [titration dose][titration unit of measure] [titration frequency] to achieve a goal [titration goal based on physiologic response, laboratory result, or assessment score]

Non-titratable IV CI:
[drug name][numerical dose][abbreviated unit strength of dose][diluent volume][volume unit of measure] "in" [compatible diluent type] "continuous intravenous infusion at" [rate][unit of measure]

Definitions:
[drug name]: The generic or brand name of the medication administered as a continuous IV infusion.
[numerical dose]: The total amount of drug contained in the prepared infusion bag or bottle, not the per-mL concentration. For example, if the order says 200 mg/100 mL, write 200 here. If the order says 1 mg/mL in 100 mL, write 100 here.
[abbreviated unit strength of dose]: The standardized abbreviated unit associated with the dose (e.g., mg, units).
[diluent volume]: The total prepared volume of the infusion bag or bottle (e.g., 100, 250). Do not write 0 when the volume is implied by a concentration such as 1 mg/mL in 100 mL.
[volume unit of measure]: The standardized abbreviated unit for the diluent volume (e.g., mL).
[compatible diluent type]: The IV fluid used to dilute the medication (e.g., 0.9% sodium chloride, D5W).
continuous intravenous infusion: The fixed route and method of administration.
[starting rate]: The initial infusion rate at which the medication is started, or the fixed infusion rate if the infusion is non-titratable (e.g., 0.05, 5).
[unit of measure]: The unit associated with the infusion rate (e.g., mcg/kg/min, units/hr, mL/hr).
[titration dose]: The numeric amount by which the infusion rate is adjusted per titration step (e.g., 0.01, 2).
[titration unit of measure]: The unit associated with the titration increment (e.g., mcg/kg/min, units/hr).
[titration frequency]: The time interval between allowable titrations (e.g., every 5 minutes, every 15 minutes).
[titration goal]: Copy the clinical target guiding titration as closely as possible from the sentence instead of paraphrasing (e.g., MAP >= 65 mmHg, RASS score -1 to 1).

Example of titratable input:
The patient was started on Ketamine continuous intravenous infusion at 0.2 mg/kg/hour using a bag of 500 mg/500 ml in 0.9% sodium chloride and titrate by 0.1 mg/kg/hour every 20 minutes to achieve a goal RASS of -4 to -5.
Example of MedMatch JSON Format:
{ "drug name": "ketamine",
"numerical dose": 500,
"abbreviated unit strength of dose": "mg",
"diluent volume": 500,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.2,
"unit of measure": "mg/kg/hr",
"titration dose": 0.1,
"titration unit of measure": "mg/kg/hr",
"titration frequency": "every 20 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5"}

Example of non-titratable input:
The patient was started on a vasopressin continuous intravenous infusion at 0.04 units/minute using a 40 units/100 ml bag in 0.9% sodium chloride.
Example of MedMatch JSON Format:
{ "drug name": "vasopressin",
"numerical dose": 40,
"abbreviated unit strength of dose": "units",
"diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.04,
"unit of measure": "units/minute"}

Example of total-dose interpretation:
The patient received a continuous intravenous infusion of midazolam at a starting rate of 0.5 mg/hr (1 mg/mL in 100 mL of 0.9% sodium chloride), and the infusion was titrated by 0.5 mg/hr every 5 minutes to achieve a RASS goal of -4 to -5.
Example of MedMatch JSON Format:
{ "drug name": "midazolam",
"numerical dose": 100,
"abbreviated unit strength of dose": "mg",
"diluent volume": 100,
"volume unit of measure": "mL",
"compatible diluent type": "0.9% sodium chloride",
"starting rate": 0.5,
"unit of measure": "mg/hr",
"titration dose": 0.5,
"titration unit of measure": "mg/hr",
"titration frequency": "every 5 minutes",
"titration goal based on physiologic response, laboratory result, or assessment score": "RASS of -4 to -5"}
"""

# ── Sheet config: sheet name -> (instruction, ground truth column mappings) ───

SHEET_CONFIG = {
    "PO Solid (40)": {
        "instruction": PO_SOLID_INSTRUCTION,
        "prompt_col": 3,   # C: Medication prompt (sentence format)
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount": 7,
            "formulation": 8,
            "route": 9,
            "frequency": 10,
        },
    },
    "PO liquid (10)": {
        "instruction": PO_LIQUID_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "numerical volume": 7,
            "volume unit of measure": 8,
            "concentration of formulation": 9,
            "formulation unit of measure": 10,
            "formulation": 11,
            "route": 12,
            "frequency": 13,
        },
    },
    "IV intermittent (16)": {
        "instruction": IV_INTERMITTENT_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "infusion time": 10,
            "frequency": 11,
        },
    },
    "IV push (17)": {
        "instruction": IV_PUSH_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of volume": 7,
            "volume unit of measure": 8,
            "concentration of solution": 9,
            "concentration unit of measure": 10,
            "formulation": 11,
            "frequency": 12,
        },
    },
    "IV continuous (16)": {
        "instruction": IV_CONTINUOUS_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "starting rate": 10,
            "unit of measure": 11,
            "titration dose": 12,
            "titration unit of measure": 13,
            "titration frequency": 14,
            "titration goal based on physiologic response, laboratory result, or assessment score": 15,
        },
    },
}

EXPECTED_KEYS_BY_SHEET = {
    sheet_name: list(config["ground_truth_cols"].keys())
    for sheet_name, config in SHEET_CONFIG.items()
}


# ── Helper functions ──────────────────────────────────────────────────────────

def load_dataset(xlsx_path):
    """Load medication data from each sheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    for sheet_name, config in SHEET_CONFIG.items():
        if SELECTED_SHEETS and sheet_name not in SELECTED_SHEETS:
            continue
        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[config["prompt_col"] - 1].value
            if not prompt:
                break
            ground_truth = {}
            for key, col_idx in config["ground_truth_cols"].items():
                val = row[col_idx - 1].value
                ground_truth[key] = val if val is not None else ""
            entries.append({
                "medication": row[0].value,
                "prompt": prompt.strip(),
                "ground_truth": ground_truth,
            })
            if MAX_ENTRIES_PER_SHEET and len(entries) >= MAX_ENTRIES_PER_SHEET:
                break
        dataset[sheet_name] = entries
        print(f"  Loaded {len(entries)} entries from '{sheet_name}'")
    return dataset


def normalize_key(key):
    """Strict: lowercase + whitespace-collapse only. No key aliases —
    rescuing wrong schema keys on the scorer side relaxes the grading standard."""
    s = str(key).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def coerce_output_object(parsed, expected_keys):
    """Convert parsed model output into a dict with the exact expected keys."""
    if parsed is None:
        return {key: "" for key in expected_keys}

    if isinstance(parsed, list):
        if len(parsed) == 1 and isinstance(parsed[0], dict):
            parsed = parsed[0]
        else:
            parsed = dict(zip(expected_keys, parsed))

    if not isinstance(parsed, dict):
        return {key: "" for key in expected_keys}

    normalized = {}
    for key, value in parsed.items():
        normalized[normalize_key(key)] = value

    return {key: normalized.get(key, "") for key in expected_keys}


def build_response_schema(expected_keys):
    """JSON schema used to force structured object output from the model."""
    return {
        "type": "object",
        "properties": {
            key: {"type": ["string", "number", "integer", "null"]}
            for key in expected_keys
        },
        "required": expected_keys,
        "additionalProperties": False,
    }


def build_instruction(sheet_name, base_instruction, prompt, expected_keys):
    """Add strict output-shape guidance and resolve continuous-infusion ambiguity."""
    extra_lines = [
        "Return one JSON object only.",
        "Do not wrap the JSON in markdown.",
        f"Use exactly these keys in this order: {', '.join(expected_keys)}.",
    ]

    if sheet_name == "IV continuous (16)":
        prompt_lower = prompt.lower()
        if any(token in prompt_lower for token in ("titrate", "titrated", "goal")):
            extra_lines.append(
                "This is a titratable continuous infusion. Use the key 'starting rate' for the initial rate."
            )
        else:
            extra_lines.append(
                "This is a non-titratable continuous infusion. Use the key 'starting rate' for the fixed infusion rate and leave titration fields as empty strings."
            )

    return f"{base_instruction}\n\n" + "\n".join(extra_lines)


def call_gemma(sheet_name, instruction, prompt, expected_keys, retries=MAX_RETRIES):
    """Send a medication prompt to the local Gemma model and return parsed JSON."""
    full_instruction = build_instruction(sheet_name, instruction, prompt, expected_keys)
    full_prompt = f"{full_instruction}\n\nNow process this medication order:\n{prompt}"

    for attempt in range(retries):
        text = ""
        try:
            text = chat_completion(
                SYSTEM_PROMPT,
                full_prompt,
                temperature=TEMPERATURE,
                model=MODEL_NAME,
            )
            # Strip markdown code fences if present
            if text.startswith("```"):
                text = text.split("\n", 1)[1] if "\n" in text else text[3:]
                if text.endswith("```"):
                    text = text[:-3].strip()
                elif "```" in text:
                    text = text[:text.rfind("```")].strip()
            # Try to parse JSON
            parsed = json.loads(text)
            return coerce_output_object(parsed, expected_keys), text
        except json.JSONDecodeError:
            # Try to extract JSON from response
            try:
                start = text.index("{")
                end = text.rindex("}") + 1
                parsed = json.loads(text[start:end])
                return coerce_output_object(parsed, expected_keys), text
            except (ValueError, json.JSONDecodeError):
                if attempt < retries - 1:
                    print(f"\n    Retry {attempt + 1}: Could not parse JSON")
                    time.sleep(RETRY_DELAY)
                else:
                    return None, text
        except Exception as e:
            if attempt < retries - 1:
                print(f"\n    Retry {attempt + 1}: API error - {e}")
                time.sleep(RETRY_DELAY)
            else:
                return None, str(e)


def normalize(value):
    """Strict: lowercase + whitespace-collapse only. No unit rewrites, no
    unicode substitution, no numeric coercion. Any local-side normalization
    beyond case/whitespace would relax the string-exact match standard."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def compare_results(llm_output, ground_truth):
    """Compare LLM output against ground truth, field by field."""
    results = {}
    for key, expected in ground_truth.items():
        expected_norm = normalize(expected)
        if llm_output and key in llm_output:
            actual_norm = normalize(llm_output[key])
        else:
            actual_norm = ""
        match = expected_norm == actual_norm
        results[key] = {
            "expected": expected_norm,
            "actual": actual_norm,
            "match": match,
        }
    return results


def all_fields_match(comparison):
    """Check if ALL fields match (per paper's accuracy definition)."""
    return all(v["match"] for v in comparison.values())


# ── Main execution ────────────────────────────────────────────────────────────

def main():
    xlsx_path = resolve_project_file("MedMatch Dataset for Experiment_ Final.xlsx")

    print(f"Model: {MODEL_NAME} | Temperature: {TEMPERATURE} | Runs: {NUM_RUNS}")
    print(f"Loading dataset from: {xlsx_path}")
    dataset = load_dataset(xlsx_path)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
    os.makedirs(results_dir, exist_ok=True)

    # Aggregate stats across all sheets and runs
    grand_total_entries = 0
    grand_total_all_correct = 0
    grand_total_fields = 0
    grand_total_fields_correct = 0

    for sheet_name, entries in dataset.items():
        config = SHEET_CONFIG[sheet_name]
        print(f"\n{'='*70}")
        print(f"Processing: {sheet_name} ({len(entries)} medications x {NUM_RUNS} runs)")
        print(f"{'='*70}")

        all_run_results = []

        for run_idx in range(NUM_RUNS):
            print(f"\n  --- Run {run_idx + 1}/{NUM_RUNS} ---")
            run_results = []
            run_all_correct = 0
            run_fields_matched = 0
            run_fields_total = 0

            for i, entry in enumerate(entries):
                print(f"  [{i+1}/{len(entries)}] {entry['medication']}...", end=" ", flush=True)

                expected_keys = EXPECTED_KEYS_BY_SHEET[sheet_name]
                llm_output, raw_response = call_gemma(
                    sheet_name,
                    config["instruction"],
                    entry["prompt"],
                    expected_keys,
                )

                comparison = compare_results(llm_output, entry["ground_truth"])
                n_match = sum(1 for v in comparison.values() if v["match"])
                n_total = len(comparison)
                entry_all_correct = all_fields_match(comparison)

                run_fields_matched += n_match
                run_fields_total += n_total
                if entry_all_correct:
                    run_all_correct += 1

                print(f"{n_match}/{n_total} fields" +
                      (" [ALL CORRECT]" if entry_all_correct else ""))

                run_results.append({
                    "run": run_idx + 1,
                    "medication": entry["medication"],
                    "prompt": entry["prompt"],
                    "ground_truth": entry["ground_truth"],
                    "llm_output": llm_output,
                    "raw_response": raw_response,
                    "comparison": comparison,
                    "fields_correct": n_match,
                    "fields_total": n_total,
                    "all_fields_correct": entry_all_correct,
                })

                # Rate limiting
                time.sleep(SLEEP_BETWEEN_CALLS)

            field_acc = (run_fields_matched / run_fields_total * 100) if run_fields_total else 0
            overall_acc = (run_all_correct / len(entries) * 100) if entries else 0
            print(f"\n  Run {run_idx+1} results:")
            print(f"    Field accuracy:   {run_fields_matched}/{run_fields_total} ({field_acc:.1f}%)")
            print(f"    Overall accuracy: {run_all_correct}/{len(entries)} ({overall_acc:.1f}%) "
                  f"(all fields correct)")

            grand_total_entries += len(entries)
            grand_total_all_correct += run_all_correct
            grand_total_fields += run_fields_total
            grand_total_fields_correct += run_fields_matched

            all_run_results.extend(run_results)

        # Save detailed JSON results for this sheet (all runs)
        safe_name = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        out_path = os.path.join(results_dir, f"{safe_name}_{timestamp}.json")
        with open(out_path, "w") as f:
            json.dump(all_run_results, f, indent=2, default=str)
        print(f"\n  JSON saved: {out_path}")

        # Save CSV summary for this sheet (all runs)
        csv_path = os.path.join(results_dir, f"{safe_name}_{timestamp}.csv")
        with open(csv_path, "w", newline="") as f:
            gt_keys = list(config["ground_truth_cols"].keys())
            writer = csv.writer(f)
            header = ["Run", "Medication"] + gt_keys + ["Fields Correct", "All Correct"]
            writer.writerow(header)
            for r in all_run_results:
                row = [r["run"], r["medication"]]
                for key in gt_keys:
                    c = r["comparison"].get(key, {})
                    if c.get("match"):
                        row.append("MATCH")
                    else:
                        row.append(f"MISS (exp: {c.get('expected','')}, got: {c.get('actual','')})")
                row.append(f"{r['fields_correct']}/{r['fields_total']}")
                row.append("YES" if r["all_fields_correct"] else "NO")
                writer.writerow(row)
        print(f"  CSV saved:  {csv_path}")

    # Grand summary
    field_pct = (grand_total_fields_correct / grand_total_fields * 100) if grand_total_fields else 0
    overall_pct = (grand_total_all_correct / grand_total_entries * 100) if grand_total_entries else 0
    print(f"\n{'='*70}")
    print(f"GRAND SUMMARY ({MODEL_NAME}, {NUM_RUNS} runs per medication)")
    print(f"  Field-level accuracy:   {grand_total_fields_correct}/{grand_total_fields} ({field_pct:.1f}%)")
    print(f"  Overall accuracy:       {grand_total_all_correct}/{grand_total_entries} ({overall_pct:.1f}%)")
    print(f"    (Overall = all fields exactly match ground truth)")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
