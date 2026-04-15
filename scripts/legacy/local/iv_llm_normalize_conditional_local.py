#!/usr/bin/env python3
"""
IV LLM-normalized experiment — CONDITIONAL variant.

Motivation: the original iv_llm_normalize_local.py applies a seven-rule
rewrite prompt on every row. On gemma4:e4b this lifts scores; on
gemma4:27b it breaks already-correct rows because the strong model
over-applies every rule and its Pass 1 output is usually already
canonical.

This variant runs deterministic flag checks on the Pass 1 output (regex
on the order sentence + the extracted JSON). For each row it builds a
Pass 2 prompt that mentions ONLY the rules whose flags fired. If no flag
fires, Pass 2 is skipped entirely and raw becomes the final answer.

This is still LLM-side normalization (the LLM rewrites the value). The
scorer is unchanged — strict case+whitespace only.

IV continuous is intentionally out of scope.
"""

import csv
import json
import os
import re
import time
from datetime import datetime

import openpyxl
from local_llm import OLLAMA_MODEL, chat_completion, resolve_project_file
from medmatch_test_local_appendix_exact import (
    IV_INTERMITTENT_INSTRUCTION,
    IV_PUSH_INSTRUCTION,
)


# ── Config ────────────────────────────────────────────────────────────────────
MODEL_NAME = os.environ.get("OLLAMA_MODEL", OLLAMA_MODEL)
TEMPERATURE = float(os.environ.get("MEDMATCH_TEMPERATURE", "0.1"))
NUM_RUNS = int(os.environ.get("MEDMATCH_NUM_RUNS", "3"))
MAX_RETRIES = int(os.environ.get("MEDMATCH_MAX_RETRIES", "3"))
RETRY_DELAY = float(os.environ.get("MEDMATCH_RETRY_DELAY", "5"))
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "0.5"))
SELECTED_SHEETS = [
    s.strip() for s in os.environ.get("MEDMATCH_SHEETS", "").split(",") if s.strip()
]
MAX_ENTRIES_PER_SHEET = int(os.environ.get("MEDMATCH_MAX_ENTRIES", "0"))

SYSTEM_PROMPT = (
    "You are a clinical pharmacist who formats medication orders. "
    "Only output the MedMatch JSON format."
)


# ── Per-rule repair text (only the ones whose flag fires are sent) ────────────
RULE_TEXT = {
    "freq": (
        "Frequency field: if the value uses an abbreviation or shorthand, rewrite it "
        "in spelled-out canonical English — BID -> twice daily, TID -> three times "
        "daily, QID -> four times daily, prn/PRN -> as needed, a bare \"daily\" -> "
        "\"once daily\", \"once a day\"/\"once per day\" -> \"once daily\", "
        "\"twice a day\" -> \"twice daily\". Preserve every qualifier verbatim "
        "(e.g., indication text, \"at bedtime\", \"for phosphorous replacement\")."
    ),
    "infusion_hyphen": (
        "Infusion time field: rewrite the hyphenated adjective form to the noun "
        "form — \"30-minute\" -> \"30 minutes\", \"3-hour\" -> \"3 hours\". "
        "Keep the numeric value unchanged."
    ),
    "ivpush_formulation": (
        "Formulation field (IV push): the current value is blank, \"vial\", or "
        "\"solution\". Set it to \"vial solution\"."
    ),
    "compound_unit": (
        "Abbreviated unit strength of dose: the order specifies a weight-based "
        "dose (e.g., mg/kg). Ensure the unit includes the \"/kg\" qualifier "
        "(or whatever weight qualifier the order uses)."
    ),
    "per_ml": (
        "Concentration fields (IV push): the source expresses drug strength as "
        "X [unit] / Y mL with Y greater than 1. Set \"concentration of solution\" "
        "to X divided by Y (as a number, e.g., 10 for 20 mg / 2 mL) and set "
        "\"concentration unit of measure\" to \"[unit]/mL\"."
    ),
    "drugname_hyphen": (
        "Drug name field: replace unicode en-dash (\u2013) or em-dash (\u2014) "
        "with an ASCII hyphen (-). Do not change drug identity or casing otherwise."
    ),
}


IV_TARGETED_PROMPT = """You are reviewing one structured IV medication order JSON for targeted format repair.

Apply ONLY the repairs listed below. Do NOT change any other fields, values, numbers, units, or wording. Keep the same JSON keys in the same order. If a listed rule does not clearly apply to a field, leave that field unchanged.

Repairs to apply:
{rules_block}

Original medication order sentence:
{sentence}

Extracted JSON to repair:
{raw_json}

Return the repaired JSON object only. No extra text, no markdown fences."""


# ── Sheet config ──────────────────────────────────────────────────────────────

SHEET_CONFIG = {
    "IV intermittent (16)": {
        "instruction": IV_INTERMITTENT_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "infusion time": 10,
            "frequency": 11,
        },
    },
    "IV push (17)": {
        "instruction": IV_PUSH_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of volume": 7,
            "volume unit of measure": 8,
            "concentration of solution": 9,
            "concentration unit of measure": 10,
            "formulation": 11,
            "frequency": 12,
        },
    },
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_value(value):
    """Strict scorer: lowercase + whitespace-collapse only."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_json_response(text):
    if text is None:
        return None
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3].strip()
        elif "```" in text:
            text = text[: text.rfind("```")].strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            start = text.index("{")
            end = text.rindex("}") + 1
            return json.loads(text[start:end])
        except (ValueError, json.JSONDecodeError):
            return None


def call_gemma(prompt_text, retries=MAX_RETRIES):
    for attempt in range(retries):
        try:
            return chat_completion(
                SYSTEM_PROMPT,
                prompt_text,
                temperature=TEMPERATURE,
                model=MODEL_NAME,
            )
        except Exception as e:
            if attempt < retries - 1:
                print(f" [retry {attempt + 1}: {e}]", end="", flush=True)
                time.sleep(RETRY_DELAY)
            else:
                return ""


def extract(instruction, prompt, expected_keys):
    """Pass 1: extract MedMatch JSON from the medication order sentence."""
    extra = (
        "Return one JSON object only.\n"
        "Do not wrap the JSON in markdown.\n"
        f"Use exactly these keys in this order: {', '.join(expected_keys)}."
    )
    full_prompt = f"{instruction}\n\n{extra}\n\nNow process this medication order:\n{prompt}"
    text = call_gemma(full_prompt)
    parsed = parse_json_response(text)
    if parsed and isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        return {k: "" for k in expected_keys}, text
    # Lowercase incoming keys so "Drug Name" maps to "drug name"
    lowered = {}
    for k, v in parsed.items():
        lk = str(k).strip().lower()
        lk = re.sub(r"\s+", " ", lk)
        lowered[lk] = v
    return {k: lowered.get(k, "") for k in expected_keys}, text


def compute_flags(sentence, raw_obj, sheet_name):
    """Return list of rule names whose flag fires for this row."""
    fired = []
    s_lower = sentence.lower()
    sheet_lower = sheet_name.lower()

    freq = str(raw_obj.get("frequency", "")).strip().lower()

    # Rule 1 — frequency canonical form
    freq_fires = False
    if re.search(r"\b(bid|tid|qid|prn)\b", freq):
        freq_fires = True
    elif re.fullmatch(r"\s*daily\s*", freq):
        freq_fires = True
    elif re.search(r"\b(once|twice)\s+(a|per)\s+day\b", freq):
        freq_fires = True
    elif re.search(r"\b(three|four)\s+times\s+(a|per)\s+day\b", freq):
        freq_fires = True
    if freq_fires:
        fired.append("freq")

    # Rule 2 — infusion time noun form (IV intermittent only)
    if "intermittent" in sheet_lower:
        inf = str(raw_obj.get("infusion time", ""))
        if re.search(r"\d+\s*-\s*(minute|hour)s?\b", inf, flags=re.IGNORECASE):
            fired.append("infusion_hyphen")

    # Rule 3 — IV push formulation default (only if the order doesn't
    # explicitly name a different injectable form).
    if "push" in sheet_lower:
        form = str(raw_obj.get("formulation", "")).strip().lower()
        if form in ("", "vial", "solution"):
            if not re.search(
                r"\b(prefilled syringe|pre-filled syringe|syringe|ampule|"
                r"ampoule|cartridge)\b",
                s_lower,
            ):
                fired.append("ivpush_formulation")

    # Rule 4 — compound unit /kg preservation
    dose_unit = str(raw_obj.get("abbreviated unit strength of dose", "")).lower()
    if "/kg" in s_lower and "/kg" not in dose_unit:
        fired.append("compound_unit")

    # Rule 5 — per-1 mL concentration reduction (IV push only)
    if "push" in sheet_lower:
        m = re.search(
            r"(\d+(?:\.\d+)?)\s*(mg|mcg|g|units?|meq)\s*/\s*(\d+(?:\.\d+)?)\s*ml",
            s_lower,
        )
        if m:
            try:
                num = float(m.group(1))
                den = float(m.group(3))
                if den > 1:
                    current = str(raw_obj.get("concentration of solution", "")).strip()
                    expected = num / den
                    try:
                        ok = abs(float(current) - expected) <= 1e-6
                    except (ValueError, TypeError):
                        ok = False
                    if not ok:
                        fired.append("per_ml")
            except (ValueError, TypeError):
                pass

    # Rule 6 — ASCII hyphens in drug name
    name = str(raw_obj.get("drug name", ""))
    if "\u2013" in name or "\u2014" in name:
        fired.append("drugname_hyphen")

    return fired


def llm_normalize_targeted(sentence, raw_obj, expected_keys, fired_rules):
    """Pass 2 with only the fired rules in-prompt. Falls back to raw if parse fails."""
    rules_block = "\n".join(f"- {RULE_TEXT[r]}" for r in fired_rules)
    raw_json_str = json.dumps(raw_obj, indent=2, default=str)
    prompt = IV_TARGETED_PROMPT.format(
        rules_block=rules_block,
        sentence=sentence,
        raw_json=raw_json_str,
    )
    text = call_gemma(prompt)
    parsed = parse_json_response(text)
    if parsed and isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        return dict(raw_obj), text
    # Lowercase keys (same as extract) so "Drug Name" still maps.
    lowered = {}
    for k, v in parsed.items():
        lk = str(k).strip().lower()
        lk = re.sub(r"\s+", " ", lk)
        lowered[lk] = v
    return {k: lowered.get(k, raw_obj.get(k, "")) for k in expected_keys}, text


def compare_results(llm_output, ground_truth):
    results = {}
    for key, expected in ground_truth.items():
        e = normalize_value(expected)
        a = normalize_value(llm_output.get(key, "") if llm_output else "")
        results[key] = {"expected": e, "actual": a, "match": e == a}
    return results


def all_fields_match(comp):
    return all(v["match"] for v in comp.values())


def load_dataset(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    for sheet_name, cfg in SHEET_CONFIG.items():
        if SELECTED_SHEETS and sheet_name not in SELECTED_SHEETS:
            continue
        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[cfg["prompt_col"] - 1].value
            if not prompt:
                break
            gt = {}
            for key, col in cfg["ground_truth_cols"].items():
                val = row[col - 1].value
                gt[key] = val if val is not None else ""
            entries.append({
                "medication": row[0].value,
                "prompt": str(prompt).strip(),
                "ground_truth": gt,
            })
            if MAX_ENTRIES_PER_SHEET and len(entries) >= MAX_ENTRIES_PER_SHEET:
                break
        dataset[sheet_name] = entries
        print(f"  Loaded {len(entries)} entries from '{sheet_name}'")
    return dataset


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    xlsx = resolve_project_file("MedMatch Dataset for Experiment_ Final.xlsx")
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
    os.makedirs(results_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"Model: {MODEL_NAME} (local Ollama) | Temp: {TEMPERATURE} | Runs: {NUM_RUNS}")
    print("Pipeline: extract -> flag check -> conditional targeted repair -> strict score")
    print("Pass 2 skipped when no flag fires.\n")
    dataset = load_dataset(xlsx)

    grand = {
        "raw_entries": 0, "raw_correct": 0, "raw_fields": 0, "raw_fields_ok": 0,
        "norm_entries": 0, "norm_correct": 0, "norm_fields": 0, "norm_fields_ok": 0,
        "pass2_calls": 0,
    }

    for sheet_name, entries in dataset.items():
        cfg = SHEET_CONFIG[sheet_name]
        expected_keys = list(cfg["ground_truth_cols"].keys())
        all_rows = []

        print(f"\n{'=' * 72}")
        print(f"{sheet_name} ({len(entries)} meds x {NUM_RUNS} runs)")
        print(f"{'=' * 72}")

        for run in range(1, NUM_RUNS + 1):
            run_raw_ok = 0
            run_norm_ok = 0
            run_pass2_calls = 0
            print(f"\n  --- Run {run}/{NUM_RUNS} ---")

            for idx, entry in enumerate(entries, 1):
                print(f"  [{idx}/{len(entries)}] {entry['medication']}...", end=" ", flush=True)

                raw_obj, raw_text = extract(cfg["instruction"], entry["prompt"], expected_keys)
                time.sleep(SLEEP_BETWEEN_CALLS)

                fired = compute_flags(entry["prompt"], raw_obj, sheet_name)
                if fired:
                    norm_obj, norm_text = llm_normalize_targeted(
                        entry["prompt"], raw_obj, expected_keys, fired,
                    )
                    time.sleep(SLEEP_BETWEEN_CALLS)
                    run_pass2_calls += 1
                else:
                    norm_obj, norm_text = dict(raw_obj), "[skipped: no flags fired]"

                comp_raw = compare_results(raw_obj, entry["ground_truth"])
                comp_norm = compare_results(norm_obj, entry["ground_truth"])
                raw_ok = all_fields_match(comp_raw)
                norm_ok = all_fields_match(comp_norm)
                raw_fc = sum(1 for v in comp_raw.values() if v["match"])
                norm_fc = sum(1 for v in comp_norm.values() if v["match"])

                flag_tag = ",".join(fired) if fired else "none"
                status = f"flags=[{flag_tag}] raw={raw_fc}/{len(expected_keys)} norm={norm_fc}/{len(expected_keys)}"
                if norm_ok and not raw_ok:
                    status += " [FIXED]"
                elif norm_ok and raw_ok:
                    status += " [ALL]"
                elif raw_ok and not norm_ok:
                    status += " [BROKE]"
                print(status)

                if raw_ok:
                    run_raw_ok += 1
                if norm_ok:
                    run_norm_ok += 1

                grand["raw_entries"] += 1
                grand["raw_correct"] += int(raw_ok)
                grand["raw_fields"] += len(expected_keys)
                grand["raw_fields_ok"] += raw_fc
                grand["norm_entries"] += 1
                grand["norm_correct"] += int(norm_ok)
                grand["norm_fields"] += len(expected_keys)
                grand["norm_fields_ok"] += norm_fc
                grand["pass2_calls"] += int(bool(fired))

                all_rows.append({
                    "run": run,
                    "medication": entry["medication"],
                    "prompt": entry["prompt"],
                    "ground_truth": entry["ground_truth"],
                    "raw_output": raw_obj,
                    "raw_response": raw_text,
                    "flags_fired": fired,
                    "normalized_output": norm_obj,
                    "normalized_response": norm_text,
                    "comparison_raw": comp_raw,
                    "comparison_normalized": comp_norm,
                    "raw_fields_correct": raw_fc,
                    "norm_fields_correct": norm_fc,
                    "raw_all_correct": raw_ok,
                    "norm_all_correct": norm_ok,
                })

            print(
                f"\n  Run {run}: raw {run_raw_ok}/{len(entries)} = "
                f"{run_raw_ok / len(entries) * 100:.1f}%, "
                f"norm {run_norm_ok}/{len(entries)} = "
                f"{run_norm_ok / len(entries) * 100:.1f}% "
                f"(pass2 calls: {run_pass2_calls}/{len(entries)})"
            )

        # Save per-sheet results
        safe = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        jpath = os.path.join(results_dir, f"{safe}_iv_llm_norm_cond_{ts}.json")
        cpath = os.path.join(results_dir, f"{safe}_iv_llm_norm_cond_{ts}.csv")

        with open(jpath, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, indent=2, default=str)

        with open(cpath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header = (
                ["Run", "Medication", "Flags"]
                + [f"raw_{k}" for k in expected_keys]
                + [f"norm_{k}" for k in expected_keys]
                + ["Raw Fields", "Norm Fields", "Raw All", "Norm All"]
            )
            writer.writerow(header)
            for r in all_rows:
                row = [r["run"], r["medication"], ",".join(r["flags_fired"]) or "none"]
                for k in expected_keys:
                    c = r["comparison_raw"][k]
                    row.append("MATCH" if c["match"]
                               else f"MISS (exp: {c['expected']}, got: {c['actual']})")
                for k in expected_keys:
                    c = r["comparison_normalized"][k]
                    row.append("MATCH" if c["match"]
                               else f"MISS (exp: {c['expected']}, got: {c['actual']})")
                row.append(f"{r['raw_fields_correct']}/{len(expected_keys)}")
                row.append(f"{r['norm_fields_correct']}/{len(expected_keys)}")
                row.append("YES" if r["raw_all_correct"] else "NO")
                row.append("YES" if r["norm_all_correct"] else "NO")
                writer.writerow(row)

        total = len(all_rows)
        raw_total_ok = sum(1 for r in all_rows if r["raw_all_correct"])
        norm_total_ok = sum(1 for r in all_rows if r["norm_all_correct"])
        raw_fc_total = sum(r["raw_fields_correct"] for r in all_rows)
        norm_fc_total = sum(r["norm_fields_correct"] for r in all_rows)
        total_fields = total * len(expected_keys)
        pass2_calls_sheet = sum(1 for r in all_rows if r["flags_fired"])
        fixed = sum(
            1 for r in all_rows
            if r["norm_all_correct"] and not r["raw_all_correct"]
        )
        broke = sum(
            1 for r in all_rows
            if r["raw_all_correct"] and not r["norm_all_correct"]
        )

        print(f"\n  {sheet_name} summary:")
        print(
            f"    Raw:  {raw_total_ok}/{total} = {raw_total_ok / total * 100:.1f}% overall, "
            f"{raw_fc_total}/{total_fields} = {raw_fc_total / total_fields * 100:.1f}% field"
        )
        print(
            f"    Norm: {norm_total_ok}/{total} = {norm_total_ok / total * 100:.1f}% overall, "
            f"{norm_fc_total}/{total_fields} = {norm_fc_total / total_fields * 100:.1f}% field"
        )
        print(f"    Pass 2 called on {pass2_calls_sheet}/{total} rows; fixed={fixed}, broke={broke}")
        print(f"  Saved: {jpath}")
        print(f"  Saved: {cpath}")

    # Grand summary
    g = grand
    print(f"\n{'=' * 72}")
    print(f"GRAND SUMMARY (IV conditional normalize, {MODEL_NAME}, {NUM_RUNS} runs)")
    if g["raw_entries"]:
        print(
            f"  Raw overall:  {g['raw_correct']}/{g['raw_entries']} = "
            f"{g['raw_correct'] / g['raw_entries'] * 100:.1f}%"
        )
        print(
            f"  Raw field:    {g['raw_fields_ok']}/{g['raw_fields']} = "
            f"{g['raw_fields_ok'] / g['raw_fields'] * 100:.1f}%"
        )
    if g["norm_entries"]:
        print(
            f"  Norm overall: {g['norm_correct']}/{g['norm_entries']} = "
            f"{g['norm_correct'] / g['norm_entries'] * 100:.1f}%"
        )
        print(
            f"  Norm field:   {g['norm_fields_ok']}/{g['norm_fields']} = "
            f"{g['norm_fields_ok'] / g['norm_fields'] * 100:.1f}%"
        )
    print(f"  Pass 2 total calls: {g['pass2_calls']}/{g['raw_entries']}")
    print(f"{'=' * 72}")


if __name__ == "__main__":
    main()
