#!/usr/bin/env python3
"""
IV LLM-normalized experiment for MedMatch (remote Google API).

Two-call pipeline for IV intermittent and IV push:
  1. Extract using the current IV instructions from medmatch_test.py
  2. Normalize the extracted JSON into canonical MedMatch wording

Scores both raw and normalized outputs with the same strict scorer used in the
local IV LLM-normalize script so the lift from the normalization pass is
directly measurable.
"""

import csv
import json
import os
import re
import sys
import time
from datetime import datetime

import openpyxl
from google import genai
from google.genai import errors, types


ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, "src")
if SRC not in sys.path:
    sys.path.insert(0, SRC)

from medmatch.core.schema import IV_INTERMITTENT_INSTRUCTION, IV_PUSH_INSTRUCTION


GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
if not GOOGLE_API_KEY and os.path.exists(os.path.join(ROOT, ".env")):
    with open(os.path.join(ROOT, ".env"), encoding="utf-8") as f:
        for line in f:
            if line.startswith("GOOGLE_API_KEY="):
                GOOGLE_API_KEY = line.strip().split("=", 1)[1]
                break

MODEL_NAME = os.environ.get("MEDMATCH_REMOTE_MODEL", "gemma-3-27b-it")
TEMPERATURE = float(os.environ.get("MEDMATCH_TEMPERATURE", "0.1"))
NUM_RUNS = int(os.environ.get("MEDMATCH_NUM_RUNS", "3"))
MAX_RETRIES = int(os.environ.get("MEDMATCH_MAX_RETRIES", "3"))
RETRY_DELAY = float(os.environ.get("MEDMATCH_RETRY_DELAY", "5"))
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "1"))
SELECTED_SHEETS = [
    s.strip() for s in os.environ.get("MEDMATCH_SHEETS", "").split(",") if s.strip()
]
MAX_ENTRIES_PER_SHEET = int(os.environ.get("MEDMATCH_MAX_ENTRIES", "0"))

CLIENT = genai.Client(api_key=GOOGLE_API_KEY)

SYSTEM_PROMPT = (
    "You are a clinical pharmacist who formats medication orders. "
    "Only output the MedMatch JSON format."
)

IV_NORMALIZE_PROMPT = """You are a clinical pharmacist reviewing a structured IV medication order JSON for formatting consistency. The JSON was extracted from a medication order sentence. Your job is to normalize the wording without changing the medical meaning.

Apply these normalization rules:

1. Frequency canonical form — write schedules in spelled-out canonical English:
   - "daily" or "once per day" -> "once daily"
   - "BID" or "twice a day" -> "twice daily"
   - "TID" -> "three times daily"
   - "QID" -> "four times daily"
   - "prn" or "PRN" -> "as needed"
   - Keep "once" as "once" when it means a one-time dose rather than a daily schedule.
   Preserve all qualifiers verbatim (e.g., "as needed for phosphorous replacement", "at bedtime", indication text). Preserve day-of-week schedules and multi-part intervals unchanged (e.g., "every 8 hours", "every 12 hours").

2. Infusion time — use the noun form "X minutes" or "X hours", not the hyphenated adjective form. Example: "30-minute" -> "30 minutes", "3-hour" -> "3 hours".

3. IV push formulation — for IV push orders, the canonical formulation is "vial solution" unless the order explicitly specifies a different injectable dosage form (e.g., "prefilled syringe"). If the extracted formulation is just "vial", just "solution", or blank, replace it with "vial solution".

4. Compound dose units — preserve qualifiers like "/kg" exactly as they appear in the order. For example, if the order says "5 mg/kg", the abbreviated unit strength of dose is "mg/kg", not "mg".

5. Concentration per 1 mL (IV push only) — if the order expresses drug strength as X [unit] / Y mL with Y greater than 1 (for example "20 mg/2 mL"), set "concentration of solution" to X divided by Y and set "concentration unit of measure" to "[unit]/mL". Example: 20 mg / 2 mL -> concentration of solution = 10, concentration unit of measure = mg/mL. If Y equals 1 (e.g., "50 mcg/mL"), the concentration is already per-mL — keep it as-is.

6. Drug name typography — use ASCII hyphens only. Replace unicode en-dash "–" or em-dash "—" inside a drug name with an ASCII hyphen "-". Example: "piperacillin–tazobactam" -> "piperacillin-tazobactam".

7. Do not change numeric values (other than rule 5), do not change drug identity, and do not add or remove fields. Return every key that appeared in the input JSON with the same key spelling.

Original medication order sentence:
{sentence}

Extracted JSON to normalize:
{raw_json}

Return the normalized JSON object only. No extra text, no markdown fences."""

SHEET_CONFIG = {
    "IV intermittent (16)": {
        "instruction": IV_INTERMITTENT_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of diluent volume": 7,
            "volume unit of measure": 8,
            "compatible diluent type": 9,
            "infusion time": 10,
            "frequency": 11,
        },
    },
    "IV push (17)": {
        "instruction": IV_PUSH_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount of volume": 7,
            "volume unit of measure": 8,
            "concentration of solution": 9,
            "concentration unit of measure": 10,
            "formulation": 11,
            "frequency": 12,
        },
    },
}


def normalize_value(value):
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_json_response(text):
    if text is None:
        return None
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3].strip()
        elif "```" in text:
            text = text[: text.rfind("```")].strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            start = text.index("{")
            end = text.rindex("}") + 1
            return json.loads(text[start:end])
        except (ValueError, json.JSONDecodeError):
            return None


def call_model(prompt_text, retries=MAX_RETRIES):
    for attempt in range(retries):
        try:
            try:
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=prompt_text,
                    config=types.GenerateContentConfig(
                        system_instruction=SYSTEM_PROMPT,
                        temperature=TEMPERATURE,
                    ),
                )
            except errors.ClientError as e:
                if "Developer instruction is not enabled" not in str(e):
                    raise
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=f"{SYSTEM_PROMPT}\n\n{prompt_text}",
                    config=types.GenerateContentConfig(temperature=TEMPERATURE),
                )
            return (response.text or "").strip()
        except Exception as e:
            if attempt < retries - 1:
                print(f" [retry {attempt + 1}: {e}]", end="", flush=True)
                time.sleep(RETRY_DELAY)
            else:
                return ""


def extract(instruction, prompt, expected_keys):
    extra = (
        "Return one JSON object only.\n"
        "Do not wrap the JSON in markdown.\n"
        f"Use exactly these keys in this order: {', '.join(expected_keys)}."
    )
    full_prompt = f"{instruction}\n\n{extra}\n\nNow process this medication order:\n{prompt}"
    text = call_model(full_prompt)
    parsed = parse_json_response(text)
    if parsed and isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        return {k: "" for k in expected_keys}, text
    return {k: parsed.get(k, "") for k in expected_keys}, text


def llm_normalize(sentence, raw_obj, expected_keys):
    raw_json_str = json.dumps(raw_obj, indent=2, default=str)
    prompt = IV_NORMALIZE_PROMPT.format(sentence=sentence, raw_json=raw_json_str)
    text = call_model(prompt)
    parsed = parse_json_response(text)
    if parsed and isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        return dict(raw_obj), text
    return {k: parsed.get(k, raw_obj.get(k, "")) for k in expected_keys}, text


def compare_results(llm_output, ground_truth):
    results = {}
    for key, expected in ground_truth.items():
        e = normalize_value(expected)
        a = normalize_value(llm_output.get(key, "") if llm_output else "")
        results[key] = {"expected": e, "actual": a, "match": e == a}
    return results


def all_fields_match(comp):
    return all(v["match"] for v in comp.values())


def load_dataset(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    for sheet_name, cfg in SHEET_CONFIG.items():
        if SELECTED_SHEETS and sheet_name not in SELECTED_SHEETS:
            continue
        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[cfg["prompt_col"] - 1].value
            if not prompt:
                break
            gt = {}
            for key, col in cfg["ground_truth_cols"].items():
                val = row[col - 1].value
                gt[key] = val if val is not None else ""
            entries.append({
                "medication": row[0].value,
                "prompt": str(prompt).strip(),
                "ground_truth": gt,
            })
            if MAX_ENTRIES_PER_SHEET and len(entries) >= MAX_ENTRIES_PER_SHEET:
                break
        dataset[sheet_name] = entries
        print(f"  Loaded {len(entries)} entries from '{sheet_name}'")
    return dataset


def main():
    xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)), "MedMatch Dataset for Experiment_ Final.xlsx")
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
    os.makedirs(results_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"Model: {MODEL_NAME} | Temp: {TEMPERATURE} | Runs: {NUM_RUNS}")
    print("Pipeline: extract -> LLM normalize -> strict score")
    print("Scoring both raw and normalized outputs to measure the lift.\n")
    dataset = load_dataset(xlsx)

    grand = {
        "raw_entries": 0, "raw_correct": 0, "raw_fields": 0, "raw_fields_ok": 0,
        "norm_entries": 0, "norm_correct": 0, "norm_fields": 0, "norm_fields_ok": 0,
    }

    for sheet_name, entries in dataset.items():
        cfg = SHEET_CONFIG[sheet_name]
        expected_keys = list(cfg["ground_truth_cols"].keys())
        all_rows = []

        print(f"\n{'='*72}")
        print(f"{sheet_name} ({len(entries)} meds x {NUM_RUNS} runs, 2 calls each)")
        print(f"{'='*72}")

        for run in range(1, NUM_RUNS + 1):
            raw_ok_count = 0
            norm_ok_count = 0
            print(f"\n  --- Run {run}/{NUM_RUNS} ---")

            for idx, entry in enumerate(entries, 1):
                print(f"  [{idx}/{len(entries)}] {entry['medication']}...", end=" ", flush=True)
                raw_obj, raw_text = extract(cfg["instruction"], entry["prompt"], expected_keys)
                time.sleep(SLEEP_BETWEEN_CALLS)
                norm_obj, norm_text = llm_normalize(entry["prompt"], raw_obj, expected_keys)
                time.sleep(SLEEP_BETWEEN_CALLS)

                comp_raw = compare_results(raw_obj, entry["ground_truth"])
                comp_norm = compare_results(norm_obj, entry["ground_truth"])
                raw_ok = all_fields_match(comp_raw)
                norm_ok = all_fields_match(comp_norm)
                raw_fc = sum(1 for v in comp_raw.values() if v["match"])
                norm_fc = sum(1 for v in comp_norm.values() if v["match"])

                status = f"raw={raw_fc}/{len(expected_keys)} norm={norm_fc}/{len(expected_keys)}"
                if norm_ok and not raw_ok:
                    status += " [FIXED]"
                elif raw_ok and not norm_ok:
                    status += " [BROKE]"
                elif norm_ok:
                    status += " [ALL]"
                print(status)

                raw_ok_count += int(raw_ok)
                norm_ok_count += int(norm_ok)
                grand["raw_entries"] += 1
                grand["raw_correct"] += int(raw_ok)
                grand["raw_fields"] += len(expected_keys)
                grand["raw_fields_ok"] += raw_fc
                grand["norm_entries"] += 1
                grand["norm_correct"] += int(norm_ok)
                grand["norm_fields"] += len(expected_keys)
                grand["norm_fields_ok"] += norm_fc

                all_rows.append({
                    "run": run,
                    "medication": entry["medication"],
                    "prompt": entry["prompt"],
                    "ground_truth": entry["ground_truth"],
                    "raw_output": raw_obj,
                    "raw_response": raw_text,
                    "normalized_output": norm_obj,
                    "normalized_response": norm_text,
                    "comparison_raw": comp_raw,
                    "comparison_normalized": comp_norm,
                    "raw_fields_correct": raw_fc,
                    "norm_fields_correct": norm_fc,
                    "raw_all_correct": raw_ok,
                    "norm_all_correct": norm_ok,
                })

            print(
                f"\n  Run {run}: raw {raw_ok_count}/{len(entries)} = {raw_ok_count/len(entries)*100:.1f}%, "
                f"normalized {norm_ok_count}/{len(entries)} = {norm_ok_count/len(entries)*100:.1f}%"
            )

        safe = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        jpath = os.path.join(results_dir, f"{safe}_iv_llm_norm_{ts}.json")
        cpath = os.path.join(results_dir, f"{safe}_iv_llm_norm_{ts}.csv")

        with open(jpath, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, indent=2, default=str)

        with open(cpath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header = (["Run", "Medication"]
                      + [f"raw_{k}" for k in expected_keys]
                      + [f"norm_{k}" for k in expected_keys]
                      + ["Raw Fields", "Norm Fields", "Raw All", "Norm All"])
            writer.writerow(header)
            for r in all_rows:
                row = [r["run"], r["medication"]]
                for k in expected_keys:
                    c = r["comparison_raw"][k]
                    row.append("MATCH" if c["match"] else f"MISS (exp: {c['expected']}, got: {c['actual']})")
                for k in expected_keys:
                    c = r["comparison_normalized"][k]
                    row.append("MATCH" if c["match"] else f"MISS (exp: {c['expected']}, got: {c['actual']})")
                row.append(f"{r['raw_fields_correct']}/{len(expected_keys)}")
                row.append(f"{r['norm_fields_correct']}/{len(expected_keys)}")
                row.append("YES" if r["raw_all_correct"] else "NO")
                row.append("YES" if r["norm_all_correct"] else "NO")
                writer.writerow(row)

        total = len(all_rows)
        raw_total_ok = sum(1 for r in all_rows if r["raw_all_correct"])
        norm_total_ok = sum(1 for r in all_rows if r["norm_all_correct"])
        raw_fc_total = sum(r["raw_fields_correct"] for r in all_rows)
        norm_fc_total = sum(r["norm_fields_correct"] for r in all_rows)
        total_fields = total * len(expected_keys)

        print(f"\n  {sheet_name} summary:")
        print(
            f"    Raw:        {raw_total_ok}/{total} = {raw_total_ok/total*100:.1f}% overall, "
            f"{raw_fc_total}/{total_fields} = {raw_fc_total/total_fields*100:.1f}% field"
        )
        print(
            f"    Normalized: {norm_total_ok}/{total} = {norm_total_ok/total*100:.1f}% overall, "
            f"{norm_fc_total}/{total_fields} = {norm_fc_total/total_fields*100:.1f}% field"
        )
        print(f"  Saved: {jpath}")
        print(f"  Saved: {cpath}")

    print(f"\n{'='*72}")
    print(f"GRAND SUMMARY (IV LLM-normalize, {MODEL_NAME}, {NUM_RUNS} runs)")
    print(f"  Raw overall: {grand['raw_correct']}/{grand['raw_entries']} = "
          f"{grand['raw_correct'] / grand['raw_entries'] * 100:.1f}%")
    print(f"  Raw field:   {grand['raw_fields_ok']}/{grand['raw_fields']} = "
          f"{grand['raw_fields_ok'] / grand['raw_fields'] * 100:.1f}%")
    print(f"  Norm overall: {grand['norm_correct']}/{grand['norm_entries']} = "
          f"{grand['norm_correct'] / grand['norm_entries'] * 100:.1f}%")
    print(f"  Norm field:   {grand['norm_fields_ok']}/{grand['norm_fields']} = "
          f"{grand['norm_fields_ok'] / grand['norm_fields'] * 100:.1f}%")
    print(f"{'='*72}")


if __name__ == "__main__":
    main()
