#!/usr/bin/env python3
"""
Oral LLM-normalized experiment for MedMatch.

Two-call pipeline:
  1. Extract: same prompt as medmatch_test.py (our refined prompts)
  2. Normalize: send the raw extraction back to the LLM with the
     ground-truth field names and formatting rules, asking it to
     normalize wording without changing meaning.

This avoids local regex hacks -- the LLM does all normalization.
Runs PO Solid and PO Liquid only.
"""

import csv
import json
import os
import re
import sys
import time
from datetime import datetime

import openpyxl
from google import genai
from google.genai import errors, types


ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, "src")
if SRC not in sys.path:
    sys.path.insert(0, SRC)

# ── Config ──────────────────────────────────────────────────────────────────
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
if not GOOGLE_API_KEY and os.path.exists(os.path.join(ROOT, ".env")):
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            if line.startswith("GOOGLE_API_KEY="):
                GOOGLE_API_KEY = line.strip().split("=", 1)[1]
                break

MODEL_NAME = "gemma-3-27b-it"
TEMPERATURE = 0.1
NUM_RUNS = int(os.environ.get("MEDMATCH_NUM_RUNS", "3"))
MAX_RETRIES = 3
RETRY_DELAY = 5
SLEEP_BETWEEN_CALLS = float(os.environ.get("MEDMATCH_SLEEP_SECONDS", "1"))

CLIENT = genai.Client(api_key=GOOGLE_API_KEY)

SYSTEM_PROMPT = (
    "You are a clinical pharmacist who formats medication orders. "
    "Only output the MedMatch JSON format."
)

# ── Extraction prompts (same as medmatch_test.py refined versions) ──────────

PO_SOLID_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for oral solid dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][amount][formulation] by mouth [frequency]

[drug name]: The generic or brand name of the medication.
[numerical dose]: The numeric amount of drug administered per dose, representing the total drug amount for that administration (e.g., 5, 10, 500). For orders with multiple identical tablets or capsules, multiply the per-unit strength by the amount (e.g., 2 capsules of 500 mg each -> numerical dose 1000, amount 2).
[abbreviated unit strength of dose]: The standardized abbreviated unit associated with the dose (e.g., mg, mcg, g).
[amount]: The number of dosage units taken per administration (e.g., 1, 2).
[formulation]: The oral solid dosage form (e.g., tablet, capsule, extended-release tablet). Copy the dosage-form wording from the order as closely as possible, including qualifiers such as extended-release or delayed-release.
by mouth: The route of administration, fixed as oral.
[frequency]: How often the medication is taken (e.g., once daily, twice daily, every 8 hours). Preserve the full schedule phrase, including qualifiers such as as needed, at bedtime, or indication text if present.

Example of input:
Gabapentin 2 capsules (total dose 600mg) by mouth three times daily.
Example of MedMatch JSON format:
{ "drug name": "gabapentin",
"numerical dose": 600,
"abbreviated unit strength of dose": "mg",
"amount": 2,
"formulation": "capsules",
"route": "by mouth",
"frequency": "three times daily"}

Example of preserving full frequency text while keeping total dose:
Give the patient a total dose of acetaminophen 1000mg by giving 2 x 500mg tablets by mouth every 8 hours as needed for pain.
Example of MedMatch JSON format:
{ "drug name": "acetaminophen",
"numerical dose": 1000,
"abbreviated unit strength of dose": "mg",
"amount": 2,
"formulation": "tablets",
"route": "by mouth",
"frequency": "every 8 hours as needed for pain"}
"""

PO_LIQUID_INSTRUCTION = """Please review the narratives about medications and format them into the MedMatch JSON format. Follow this exact slot order; if a slot is unknown, use an empty string and do not fabricate.

The MedMatch JSON format for oral liquid dosage form medications is:
[drug name][numerical dose][abbreviated unit strength of dose][numerical volume][abbreviated unit strength of volume] of the [concentration of formulation][formulation unit of measure][formulation] by mouth [frequency]

[drug name]: The generic or brand name of the medication.
[numerical dose]: The numeric amount of drug administered per dose (e.g., 250, 5).
[abbreviated unit strength of dose]: The standardized abbreviated unit for the drug dose (e.g., mg, mcg).
[numerical volume]: The numeric volume administered per dose (e.g., 5, 10).
[abbreviated unit strength of volume]: The standardized abbreviated unit for volume (e.g., mL).
[concentration of formulation]: The strength of the medication per 1 mL. For example, if the sentence says 250 mg/5 mL, write 50 here.
[formulation unit of measure]: The unit used in the concentration denominator (e.g., mL).
[formulation]: The oral liquid dosage form (e.g., solution, suspension, syrup).
[route]: The route of administration, fixed as oral.
[frequency]: How often the medication is administered (e.g., once daily, every 6 hours).

Example of input:
An oral solution of diazepam (1mg/mL) is used to administer 5mg (5mL) dose by mouth once daily.
Example of MedMatch JSON Format:
{ "drug name": "diazepam",
"numerical dose": 5,
"abbreviated unit strength of dose": "mg",
"numerical volume": 5,
"volume unit of measure": "mL",
"concentration of formulation": 1,
"formulation unit of measure": "mg/mL",
"formulation": "solution",
"route": "by mouth",
"frequency": "once daily"}

Example of concentration conversion:
Using a 250mg/5mL solution, administer a dose of 500mg of valproic acid (10mL) every 6 hours by mouth.
Example of MedMatch JSON Format:
{ "drug name": "valproic acid",
"numerical dose": 500,
"abbreviated unit strength of dose": "mg",
"numerical volume": 10,
"volume unit of measure": "mL",
"concentration of formulation": 50,
"formulation unit of measure": "mg/mL",
"formulation": "solution",
"route": "by mouth",
"frequency": "every 6 hours"}
"""

# ── Normalization prompt (second LLM call) ──────────────────────────────────

NORMALIZE_PROMPT = """You are a clinical pharmacist reviewing a structured medication order JSON for formatting consistency. The JSON was extracted from a medication order sentence. Your job is to normalize the wording without changing the medical meaning.

Apply these normalization rules:
1. Frequency: If the schedule means once per day, write "once daily" (not just "daily"). Preserve all qualifiers like "as needed", "at bedtime", indication text (e.g., "as needed for pain"), and day-of-week schedules. Keep the order: frequency first, then qualifiers (e.g., "every 4 hours as needed for pain", not "as needed for pain every 4 hours"). If the order specifies specific days, write as "once daily on [days]".
2. Formulation: Use hyphens for multi-word dosage forms: "extended-release tablet" not "extended release tablet". Use singular form (e.g., "tablet" not "tablets", "capsule" not "capsules") unless the amount field is greater than 1, in which case use plural.
3. Route: Always write "by mouth" (not "oral" or "po").
4. Units: Use standard abbreviations: mg, mcg, g, mL, mg/mL.
5. Oral solid total dose: For oral solid orders only, if the source sentence explicitly states a total dose or gives multiple identical tablets/capsules with a per-unit strength, set numerical dose to the total drug amount per administration and keep amount as the dosage-unit count. Example: 2 capsules of 500 mg each -> numerical dose 1000, amount 2. Do not infer missing strengths or counts.
6. Otherwise, do not change numeric values, drug names, or any medical content. Only fix formatting and wording.

Original medication order sentence:
{sentence}

Extracted JSON to normalize:
{raw_json}

Return the normalized JSON only. No extra text."""

# ── Sheet config ────────────────────────────────────────────────────────────

SHEET_CONFIG = {
    "PO Solid (40)": {
        "instruction": PO_SOLID_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "amount": 7,
            "formulation": 8,
            "route": 9,
            "frequency": 10,
        },
    },
    "PO liquid (10)": {
        "instruction": PO_LIQUID_INSTRUCTION,
        "prompt_col": 3,
        "ground_truth_cols": {
            "drug name": 4,
            "numerical dose": 5,
            "abbreviated unit strength of dose": 6,
            "numerical volume": 7,
            "volume unit of measure": 8,
            "concentration of formulation": 9,
            "formulation unit of measure": 10,
            "formulation": 11,
            "route": 12,
            "frequency": 13,
        },
    },
}


# ── Helpers ─────────────────────────────────────────────────────────────────

def normalize_value(value):
    """Normalize a value for comparison (same as medmatch_test.py)."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = s.replace("–", "-").replace("—", "-").replace("≥", ">=")
    s = re.sub(r"\bmg/hour\b", "mg/hr", s)
    s = re.sub(r"\bmcg/hour\b", "mcg/hr", s)
    s = re.sub(r"\s+", " ", s)
    try:
        f = float(s)
        s = str(int(f)) if f == int(f) else str(f)
    except (ValueError, OverflowError):
        pass
    return s


def parse_json_response(text):
    """Parse JSON from model response, handling markdown fences."""
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3].strip()
        elif "```" in text:
            text = text[:text.rfind("```")].strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            start = text.index("{")
            end = text.rindex("}") + 1
            return json.loads(text[start:end])
        except (ValueError, json.JSONDecodeError):
            return None


def call_gemma(prompt_text, retries=MAX_RETRIES):
    """Send a prompt to Gemma and return raw text."""
    for attempt in range(retries):
        try:
            try:
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=prompt_text,
                    config=types.GenerateContentConfig(
                        system_instruction=SYSTEM_PROMPT,
                        temperature=TEMPERATURE,
                    ),
                )
            except errors.ClientError as e:
                if "Developer instruction is not enabled" not in str(e):
                    raise
                response = CLIENT.models.generate_content(
                    model=MODEL_NAME,
                    contents=f"{SYSTEM_PROMPT}\n\n{prompt_text}",
                    config=types.GenerateContentConfig(temperature=TEMPERATURE),
                )
            return response.text.strip()
        except Exception as e:
            if attempt < retries - 1:
                print(f" [retry {attempt+1}: {e}]", end="", flush=True)
                time.sleep(RETRY_DELAY)
            else:
                return ""


def extract(instruction, prompt, expected_keys):
    """Call 1: extract MedMatch JSON from sentence."""
    extra = (
        "Return one JSON object only.\n"
        "Do not wrap the JSON in markdown.\n"
        f"Use exactly these keys in this order: {', '.join(expected_keys)}."
    )
    full_prompt = f"{instruction}\n\n{extra}\n\nNow process this medication order:\n{prompt}"
    text = call_gemma(full_prompt)
    parsed = parse_json_response(text)
    if parsed and isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        return {k: "" for k in expected_keys}, text
    return {k: parsed.get(k, "") for k in expected_keys}, text


def llm_normalize(sentence, raw_obj, expected_keys):
    """Call 2: ask LLM to normalize the extracted JSON."""
    raw_json_str = json.dumps(raw_obj, indent=2)
    prompt = NORMALIZE_PROMPT.format(sentence=sentence, raw_json=raw_json_str)
    text = call_gemma(prompt)
    parsed = parse_json_response(text)
    if not isinstance(parsed, dict):
        return raw_obj, text  # fallback to raw if normalization fails
    return {k: parsed.get(k, raw_obj.get(k, "")) for k in expected_keys}, text


def compare_results(llm_output, ground_truth):
    results = {}
    for key, expected in ground_truth.items():
        e = normalize_value(expected)
        a = normalize_value(llm_output.get(key, "") if llm_output else "")
        results[key] = {"expected": e, "actual": a, "match": e == a}
    return results


def all_fields_match(comp):
    return all(v["match"] for v in comp.values())


def load_dataset(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    for sheet_name, cfg in SHEET_CONFIG.items():
        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prompt = row[cfg["prompt_col"] - 1].value
            if not prompt:
                break
            gt = {}
            for key, col in cfg["ground_truth_cols"].items():
                val = row[col - 1].value
                gt[key] = val if val is not None else ""
            entries.append({
                "medication": row[0].value,
                "prompt": str(prompt).strip(),
                "ground_truth": gt,
            })
        dataset[sheet_name] = entries
        print(f"  Loaded {len(entries)} from {sheet_name}")
    return dataset


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "MedMatch Dataset for Experiment_ Final.xlsx")
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
    os.makedirs(results_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"Model: {MODEL_NAME} | Temp: {TEMPERATURE} | Runs: {NUM_RUNS}")
    print(f"Pipeline: extract -> LLM normalize -> score\n")
    dataset = load_dataset(xlsx)

    for sheet_name, entries in dataset.items():
        cfg = SHEET_CONFIG[sheet_name]
        expected_keys = list(cfg["ground_truth_cols"].keys())
        all_rows = []

        print(f"\n{'='*60}")
        print(f"{sheet_name} ({len(entries)} meds x {NUM_RUNS} runs, 2 calls each)")
        print(f"{'='*60}")

        for run in range(1, NUM_RUNS + 1):
            run_ok = 0
            print(f"\n  --- Run {run}/{NUM_RUNS} ---")

            for idx, entry in enumerate(entries, 1):
                print(f"  [{idx}/{len(entries)}] {entry['medication']}...", end=" ", flush=True)

                # Call 1: extract
                raw_obj, raw_text = extract(cfg["instruction"], entry["prompt"], expected_keys)
                time.sleep(SLEEP_BETWEEN_CALLS)

                # Call 2: LLM normalize
                norm_obj, norm_text = llm_normalize(entry["prompt"], raw_obj, expected_keys)
                time.sleep(SLEEP_BETWEEN_CALLS)

                # Score against ground truth
                comp_raw = compare_results(raw_obj, entry["ground_truth"])
                comp_norm = compare_results(norm_obj, entry["ground_truth"])
                raw_ok = all_fields_match(comp_raw)
                norm_ok = all_fields_match(comp_norm)

                raw_fc = sum(1 for v in comp_raw.values() if v["match"])
                norm_fc = sum(1 for v in comp_norm.values() if v["match"])

                status = f"raw={raw_fc}/{len(expected_keys)} norm={norm_fc}/{len(expected_keys)}"
                if norm_ok and not raw_ok:
                    status += " [FIXED]"
                elif norm_ok:
                    status += " [ALL]"
                print(status)

                if norm_ok:
                    run_ok += 1

                all_rows.append({
                    "run": run,
                    "medication": entry["medication"],
                    "prompt": entry["prompt"],
                    "ground_truth": entry["ground_truth"],
                    "raw_output": raw_obj,
                    "raw_response": raw_text,
                    "normalized_output": norm_obj,
                    "normalized_response": norm_text,
                    "comparison_raw": comp_raw,
                    "comparison_normalized": comp_norm,
                    "raw_fields_correct": raw_fc,
                    "norm_fields_correct": norm_fc,
                    "raw_all_correct": raw_ok,
                    "norm_all_correct": norm_ok,
                })

            print(f"\n  Run {run}: {run_ok}/{len(entries)} = {run_ok/len(entries)*100:.1f}%")

        # Save results
        safe = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        jpath = os.path.join(results_dir, f"{safe}_llm_norm_{ts}.json")
        cpath = os.path.join(results_dir, f"{safe}_llm_norm_{ts}.csv")

        with open(jpath, "w") as f:
            json.dump(all_rows, f, indent=2, default=str)

        with open(cpath, "w", newline="") as f:
            writer = csv.writer(f)
            header = (["Run", "Medication"]
                      + [f"raw_{k}" for k in expected_keys]
                      + [f"norm_{k}" for k in expected_keys]
                      + ["Raw Fields", "Norm Fields", "Raw All", "Norm All"])
            writer.writerow(header)
            for r in all_rows:
                row = [r["run"], r["medication"]]
                for k in expected_keys:
                    c = r["comparison_raw"][k]
                    row.append("MATCH" if c["match"] else f"MISS (exp: {c['expected']}, got: {c['actual']})")
                for k in expected_keys:
                    c = r["comparison_normalized"][k]
                    row.append("MATCH" if c["match"] else f"MISS (exp: {c['expected']}, got: {c['actual']})")
                row.append(f"{r['raw_fields_correct']}/{len(expected_keys)}")
                row.append(f"{r['norm_fields_correct']}/{len(expected_keys)}")
                row.append("YES" if r["raw_all_correct"] else "NO")
                row.append("YES" if r["norm_all_correct"] else "NO")
                writer.writerow(row)

        # Summary
        total = len(all_rows)
        raw_total_ok = sum(1 for r in all_rows if r["raw_all_correct"])
        norm_total_ok = sum(1 for r in all_rows if r["norm_all_correct"])
        raw_fc_total = sum(r["raw_fields_correct"] for r in all_rows)
        norm_fc_total = sum(r["norm_fields_correct"] for r in all_rows)
        total_fields = total * len(expected_keys)

        print(f"\n  {sheet_name} summary:")
        print(f"    Raw:        {raw_total_ok}/{total} = {raw_total_ok/total*100:.1f}% overall, "
              f"{raw_fc_total}/{total_fields} = {raw_fc_total/total_fields*100:.1f}% field")
        print(f"    Normalized: {norm_total_ok}/{total} = {norm_total_ok/total*100:.1f}% overall, "
              f"{norm_fc_total}/{total_fields} = {norm_fc_total/total_fields*100:.1f}% field")
        print(f"  Saved: {cpath}")


if __name__ == "__main__":
    main()
