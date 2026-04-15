"""Dataset loading helpers for workbook-backed MedMatch experiments."""

import os

import openpyxl


def resolve_project_file(filename, start_dir=None):
    base_dir = os.path.abspath(start_dir or os.getcwd())
    candidates = [
        os.path.join(base_dir, "datasets", filename),
        os.path.join(os.path.dirname(base_dir), "datasets", filename),
        os.path.join(base_dir, filename),
        os.path.join(os.path.dirname(base_dir), filename),
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return candidates[-1]


def selected_sheets_from_env(default=None):
    raw = os.environ.get("MEDMATCH_SHEETS", "")
    selected = [sheet.strip() for sheet in raw.split(",") if sheet.strip()]
    return selected or list(default or [])


def load_dataset(xlsx_path, sheet_config, selected_sheets=None, max_entries_per_sheet=0):
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    dataset = {}
    selected = set(selected_sheets or sheet_config.keys())

    for sheet_name, config in sheet_config.items():
        if sheet_name not in selected:
            continue
        worksheet = workbook[sheet_name]
        rows = []
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            prompt = row[config["prompt_col"] - 1].value
            if not prompt:
                break
            ground_truth = {
                key: (row[col_idx - 1].value if row[col_idx - 1].value is not None else "")
                for key, col_idx in config["ground_truth_cols"].items()
            }
            rows.append(
                {
                    "medication": row[0].value,
                    "prompt": str(prompt).strip(),
                    "ground_truth": ground_truth,
                }
            )
            if max_entries_per_sheet and len(rows) >= max_entries_per_sheet:
                break
        dataset[sheet_name] = rows
    return dataset
